import sys
from PySide6.QtWidgets import (QApplication, QMainWindow, QTabWidget, QWidget, 
                               QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem,
                               QPushButton, QDateEdit, QLabel, QFileDialog, QMessageBox,
                               QHeaderView, QComboBox, QTimeEdit, QTextEdit, QDialog,
                               QFormLayout, QDialogButtonBox, QGroupBox, QRadioButton,
                               QSpinBox, QSplitter, QLineEdit, QCalendarWidget, QGridLayout,
                               QFrame, QScrollArea, QProgressBar)
from PySide6.QtCore import Qt, QDate, QTime, QLocale, Signal
from PySide6.QtGui import QFont, QTextCharFormat, QColor
from datetime import datetime, date, timedelta
import traceback

from database import DatabaseManager
from main import ExcelProcessor
from database_utils import check_database_status, force_unlock_database, diagnose_database_lock
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class IndonesianCalendar(QCalendarWidget):
    """Kalender custom dengan bahasa Indonesia dan tanggal merah untuk hari Minggu"""
    def __init__(self, parent=None):
        super().__init__(parent)
        
        # Set locale ke Indonesia
        locale = QLocale(QLocale.Indonesian, QLocale.Indonesia)
        self.setLocale(locale)
        
        # Nama hari dalam bahasa Indonesia
        self.setHorizontalHeaderFormat(QCalendarWidget.LongDayNames)
        
        # Format untuk hari Minggu (tanggal merah)
        sunday_format = QTextCharFormat()
        sunday_format.setForeground(QColor(255, 0, 0))  # Merah
        sunday_format.setBackground(QColor(255, 230, 230))  # Background pink muda
        self.setWeekdayTextFormat(Qt.Sunday, sunday_format)

class IndonesianDateEdit(QDateEdit):
    """DateEdit dengan kalender bahasa Indonesia dan tanggal merah"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setCalendarPopup(True)
        
        # Buat custom calendar popup
        calendar = IndonesianCalendar(self)
        self.setCalendarWidget(calendar)
        
        # Set format tampilan tanggal
        self.setDisplayFormat("dd MMMM yyyy")
        
        # Set locale ke Indonesia untuk format tanggal
        locale = QLocale(QLocale.Indonesian, QLocale.Indonesia)
        self.setLocale(locale)

class ViolationDialog(QDialog):
    def __init__(self, employees, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Tambah Pelanggaran")
        self.setModal(True)
        self.resize(400, 300)
        
        layout = QFormLayout()
        
        # Dropdown karyawan
        self.employee_combo = QComboBox()
        for emp in employees:
            self.employee_combo.addItem(emp['Nama'], emp.get('id'))
        layout.addRow("Karyawan:", self.employee_combo)
        
        # Waktu mulai dan selesai dengan detik
        self.start_time = QTimeEdit()
        self.start_time.setDisplayFormat("HH:mm:ss")
        layout.addRow("Jam Mulai:", self.start_time)
        
        self.end_time = QTimeEdit()
        self.end_time.setDisplayFormat("HH:mm:ss")
        layout.addRow("Jam Selesai:", self.end_time)
        
        # Keterangan
        self.description = QTextEdit()
        self.description.setMaximumHeight(100)
        layout.addRow("Keterangan:", self.description)
        
        # Buttons
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addRow(buttons)
        
        self.setLayout(layout)
    
    def get_violation_data(self):
        return {
            'employee_name': self.employee_combo.currentText(),
            'attendance_id': self.employee_combo.currentData(),
            'start_time': self.start_time.time().toString("HH:mm:ss"),
            'end_time': self.end_time.time().toString("HH:mm:ss"),
            'description': self.description.toPlainText()
        }

class LeaveManagementDialog(QDialog):
    def __init__(self, db_manager, employee_id, employee_name, date, parent=None):
        super().__init__(parent)
        self.db_manager = db_manager
        self.employee_id = employee_id
        self.employee_name = employee_name
        self.date = date
        self.setWindowTitle(f"Kelola Izin - {employee_name} ({date})")
        self.setModal(True)
        self.resize(800, 500)
        
        self.init_ui()
        self.load_leaves()
    
    def init_ui(self):
        layout = QVBoxLayout()
        
        # Header info
        info_label = QLabel(f"Kelola Izin untuk: {self.employee_name}\nTanggal: {self.date}")
        info_label.setStyleSheet("font-weight: bold; padding: 10px; background-color: #e8f5e8; border-radius: 5px;")
        layout.addWidget(info_label)
        
        # Buttons
        btn_layout = QHBoxLayout()
        
        self.add_btn = QPushButton("➕ Tambah Izin")
        self.add_btn.setStyleSheet("""
            QPushButton {
                background-color: #28a745;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px 15px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #218838;
            }
        """)
        self.add_btn.clicked.connect(self.add_leave)
        btn_layout.addWidget(self.add_btn)
        
        self.edit_btn = QPushButton("✏️ Edit Izin")
        self.edit_btn.setStyleSheet("""
            QPushButton {
                background-color: #ffc107;
                color: black;
                border: none;
                border-radius: 5px;
                padding: 8px 15px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #e0a800;
            }
        """)
        self.edit_btn.clicked.connect(self.edit_leave)
        self.edit_btn.setEnabled(False)
        btn_layout.addWidget(self.edit_btn)
        
        self.delete_btn = QPushButton("🗑️ Hapus Izin")
        self.delete_btn.setStyleSheet("""
            QPushButton {
                background-color: #dc3545;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px 15px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #c82333;
            }
        """)
        self.delete_btn.clicked.connect(self.delete_leave)
        self.delete_btn.setEnabled(False)
        btn_layout.addWidget(self.delete_btn)
        
        btn_layout.addStretch()
        layout.addLayout(btn_layout)
        
        # Table
        self.table = QTableWidget()
        self.table.setColumnCount(3)
        self.table.setHorizontalHeaderLabels(["Keterangan", "Dibuat", "ID"])
        
        # Hide ID column
        self.table.setColumnHidden(2, True)
        
        # Resize columns
        header = self.table.horizontalHeader()
        for i in range(2):  # Exclude hidden ID column
            header.setSectionResizeMode(i, QHeaderView.Interactive)
        
        # Set default width untuk kolom
        self.table.setColumnWidth(0, 400)  # Keterangan
        self.table.setColumnWidth(1, 150)  # Dibuat
        
        # Enable stretching table to fill available space
        self.table.horizontalHeader().setStretchLastSection(True)
        
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.selectionModel().selectionChanged.connect(self.on_selection_changed)
        
        # Set row height
        self.table.verticalHeader().setDefaultSectionSize(45)
        
        layout.addWidget(self.table)
        
        # Close button
        close_layout = QHBoxLayout()
        close_layout.addStretch()
        close_btn = QPushButton("Tutup")
        close_btn.clicked.connect(self.accept)
        close_layout.addWidget(close_btn)
        layout.addLayout(close_layout)
        
        self.setLayout(layout)
    
    def load_leaves(self):
        """Load leaves data into table"""
        try:
            leaves = self.db_manager.get_leaves_by_employee_date(self.employee_id, self.date)
            self.table.setRowCount(len(leaves))
            
            for row, leave in enumerate(leaves):
                # Keterangan
                self.table.setItem(row, 0, QTableWidgetItem(leave['description']))
                
                # Format created_at timestamp
                created_at = leave['created_at']
                if created_at:
                    from datetime import datetime
                    try:
                        dt = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
                        formatted_date = dt.strftime("%d/%m/%Y %H:%M")
                    except:
                        formatted_date = created_at
                else:
                    formatted_date = "-"
                
                self.table.setItem(row, 1, QTableWidgetItem(formatted_date))
                
                # Store leave ID in hidden column
                id_item = QTableWidgetItem(str(leave['id']))
                self.table.setItem(row, 2, id_item)
                
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal memuat data izin:\n{str(e)}")
    
    def on_selection_changed(self):
        """Handle table selection change"""
        has_selection = len(self.table.selectionModel().selectedRows()) > 0
        self.edit_btn.setEnabled(has_selection)
        self.delete_btn.setEnabled(has_selection)
    
    def add_leave(self):
        """Add new leave"""
        dialog = LeaveDialog(self.db_manager, self.employee_id, self.employee_name, self.date, self)
        if dialog.exec() == QDialog.Accepted:
            self.load_leaves()
    
    def edit_leave(self):
        """Edit selected leave"""
        selected_rows = self.table.selectionModel().selectedRows()
        if not selected_rows:
            return
        
        row = selected_rows[0].row()
        leave_id = int(self.table.item(row, 2).text())  # Get ID from hidden column
        description = self.table.item(row, 0).text()
        
        dialog = LeaveDialog(self.db_manager, self.employee_id, self.employee_name, self.date, self, 
                           leave_id=leave_id, description=description)
        if dialog.exec() == QDialog.Accepted:
            self.load_leaves()
    
    def delete_leave(self):
        """Delete selected leave"""
        selected_rows = self.table.selectionModel().selectedRows()
        if not selected_rows:
            return
        
        row = selected_rows[0].row()
        leave_id = int(self.table.item(row, 2).text())
        description = self.table.item(row, 0).text()
        
        reply = QMessageBox.question(self, "Konfirmasi Hapus", 
                                   f"Apakah Anda yakin ingin menghapus izin:\n\"{description}\"?",
                                   QMessageBox.Yes | QMessageBox.No)
        
        if reply == QMessageBox.Yes:
            try:
                self.db_manager.delete_leave(leave_id)
                self.load_leaves()
                QMessageBox.information(self, "Success", "Izin berhasil dihapus!")
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Gagal menghapus izin:\n{str(e)}")


class LeaveSelectionDialog(QDialog):
    def __init__(self, employees, parent=None):
        super().__init__(parent)
        self.employees = employees
        self.setWindowTitle("Pilih Karyawan untuk Izin")
        self.setModal(True)
        self.resize(400, 200)
        
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        
        # Header
        header = QLabel("Pilih Karyawan untuk Menambahkan Izin")
        header.setStyleSheet("font-size: 14px; font-weight: bold; padding: 10px; color: #28a745;")
        header.setAlignment(Qt.AlignCenter)
        layout.addWidget(header)
        
        # Employee selection
        form_layout = QFormLayout()
        
        self.employee_combo = QComboBox()
        self.employee_combo.addItems(self.employees)
        form_layout.addRow("Karyawan:", self.employee_combo)
        
        layout.addLayout(form_layout)
        
        # Buttons
        btn_layout = QHBoxLayout()
        
        ok_btn = QPushButton("OK")
        ok_btn.setStyleSheet("""
            QPushButton {
                background-color: #28a745;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px 20px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #218838;
            }
        """)
        ok_btn.clicked.connect(self.accept)
        btn_layout.addWidget(ok_btn)
        
        cancel_btn = QPushButton("Batal")
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(cancel_btn)
        
        layout.addLayout(btn_layout)
        self.setLayout(layout)
    
    def get_selected_employee(self):
        return self.employee_combo.currentText()


class LeaveDialog(QDialog):
    def __init__(self, db_manager, employee_id, employee_name, date, parent=None, leave_id=None, description=""):
        super().__init__(parent)
        self.db_manager = db_manager
        self.employee_id = employee_id
        self.employee_name = employee_name
        self.date = date
        self.leave_id = leave_id
        self.is_edit = leave_id is not None
        
        self.setWindowTitle("Edit Izin" if self.is_edit else "Tambah Izin")
        self.setModal(True)
        self.resize(500, 300)
        
        self.init_ui()
        
        # Fill data if editing
        if self.is_edit:
            self.description_edit.setText(description)
    
    def init_ui(self):
        layout = QVBoxLayout()
        
        # Header
        header = QLabel("Edit Izin" if self.is_edit else "Tambah Izin Baru")
        header.setStyleSheet("font-size: 16px; font-weight: bold; padding: 10px; color: #28a745;")
        header.setAlignment(Qt.AlignCenter)
        layout.addWidget(header)
        
        # Form
        form_layout = QFormLayout()
        
        # Employee (read-only)
        employee_label = QLabel(self.employee_name)
        employee_label.setStyleSheet("padding: 8px; background-color: #f8f9fa; border: 1px solid #dee2e6; border-radius: 4px;")
        form_layout.addRow("Karyawan:", employee_label)
        
        # Date (read-only)
        date_label = QLabel(self.date)
        date_label.setStyleSheet("padding: 8px; background-color: #f8f9fa; border: 1px solid #dee2e6; border-radius: 4px;")
        form_layout.addRow("Tanggal:", date_label)
        
        # Description
        self.description_edit = QTextEdit()
        self.description_edit.setMaximumHeight(100)
        self.description_edit.setPlaceholderText("Masukkan keterangan izin (contoh: Sakit, Urusan keluarga, dll)")
        form_layout.addRow("Keterangan:", self.description_edit)
        
        layout.addLayout(form_layout)
        
        # Buttons
        btn_layout = QHBoxLayout()
        
        save_btn = QPushButton("💾 Simpan")
        save_btn.setStyleSheet("""
            QPushButton {
                background-color: #28a745;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px 20px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #218838;
            }
        """)
        save_btn.clicked.connect(self.save_leave)
        btn_layout.addWidget(save_btn)
        
        cancel_btn = QPushButton("❌ Batal")
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(cancel_btn)
        
        layout.addLayout(btn_layout)
        self.setLayout(layout)
    
    def save_leave(self):
        """Save leave data"""
        description = self.description_edit.toPlainText().strip()
        
        if not description:
            QMessageBox.warning(self, "Warning", "Keterangan izin tidak boleh kosong!")
            return
        
        try:
            if self.is_edit:
                self.db_manager.update_leave(self.leave_id, self.employee_id, self.date, description)
                QMessageBox.information(self, "Success", "Izin berhasil diupdate!")
            else:
                self.db_manager.add_leave(self.employee_id, self.date, description)
                QMessageBox.information(self, "Success", "Izin berhasil ditambahkan!")
            
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal menyimpan izin:\n{str(e)}")


class ViolationManagementDialog(QDialog):
    def __init__(self, db_manager, attendance_id, employee_name, parent=None):
        super().__init__(parent)
        self.db_manager = db_manager
        self.attendance_id = attendance_id
        self.employee_name = employee_name
        
        self.setWindowTitle(f"Kelola Pelanggaran - {employee_name}")
        self.setModal(True)
        self.resize(800, 500)
        self.init_ui()
        self.load_violations()
    
    def init_ui(self):
        layout = QVBoxLayout()
        
        # Header info
        info_label = QLabel(f"Kelola Pelanggaran untuk: {self.employee_name}")
        info_label.setStyleSheet("font-weight: bold; font-size: 14px; padding: 10px;")
        layout.addWidget(info_label)
        
        # Buttons
        button_layout = QHBoxLayout()
        
        self.add_btn = QPushButton("➕ Tambah Pelanggaran")
        self.add_btn.clicked.connect(self.add_violation)
        button_layout.addWidget(self.add_btn)
        
        self.edit_btn = QPushButton("✏️ Edit Pelanggaran")
        self.edit_btn.clicked.connect(self.edit_violation)
        self.edit_btn.setEnabled(False)
        button_layout.addWidget(self.edit_btn)
        
        self.delete_btn = QPushButton("🗑️ Hapus Pelanggaran")
        self.delete_btn.clicked.connect(self.delete_violation)
        self.delete_btn.setEnabled(False)
        button_layout.addWidget(self.delete_btn)
        
        button_layout.addStretch()
        layout.addLayout(button_layout)
        
        # Table
        self.table = QTableWidget()
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(["Jam Mulai", "Jam Selesai", "Keterangan", "Dibuat"])
        
        # Table settings - ubah ke Interactive agar pengguna dapat mengubah ukuran kolom
        header = self.table.horizontalHeader()
        
        # Set semua kolom ke Interactive (bisa diubah ukurannya oleh user)
        for i in range(4):  # Semua kolom
            header.setSectionResizeMode(i, QHeaderView.Interactive)
        
        # Set default width untuk kolom
        self.table.setColumnWidth(0, 100)  # Jam Mulai
        self.table.setColumnWidth(1, 100)  # Jam Selesai
        self.table.setColumnWidth(2, 400)  # Keterangan
        self.table.setColumnWidth(3, 150)  # Dibuat
        
        # Enable stretching table to fill available space
        self.table.horizontalHeader().setStretchLastSection(True)
        
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.selectionModel().selectionChanged.connect(self.on_selection_changed)
        layout.addWidget(self.table)
        
        # Close button
        close_layout = QHBoxLayout()
        close_layout.addStretch()
        close_btn = QPushButton("Tutup")
        close_btn.clicked.connect(self.accept)
        close_layout.addWidget(close_btn)
        layout.addLayout(close_layout)
        
        self.setLayout(layout)
    
    def load_violations(self):
        """Load violations data into table"""
        try:
            violations = self.db_manager.get_violations_by_attendance(self.attendance_id)
            self.table.setRowCount(len(violations))
            
            for row, violation in enumerate(violations):
                # Store violation ID in first column as hidden data
                start_item = QTableWidgetItem(violation['start_time'])
                start_item.setData(Qt.UserRole, violation['id'])  # Store violation ID
                self.table.setItem(row, 0, start_item)
                
                self.table.setItem(row, 1, QTableWidgetItem(violation['end_time']))
                self.table.setItem(row, 2, QTableWidgetItem(violation['description']))
                
                # Format created_at timestamp
                created_at = violation['created_at']
                if created_at:
                    from datetime import datetime
                    try:
                        dt = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
                        formatted_date = dt.strftime("%d/%m/%Y %H:%M")
                    except:
                        formatted_date = created_at
                else:
                    formatted_date = "-"
                
                self.table.setItem(row, 3, QTableWidgetItem(formatted_date))
                
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal memuat data pelanggaran:\n{str(e)}")
    
    def on_selection_changed(self):
        """Handle table selection change"""
        has_selection = len(self.table.selectionModel().selectedRows()) > 0
        self.edit_btn.setEnabled(has_selection)
        self.delete_btn.setEnabled(has_selection)
    
    def add_violation(self):
        """Add new violation"""
        dialog = ViolationEditDialog(parent=self)
        if dialog.exec() == QDialog.Accepted:
            violation_data = dialog.get_violation_data()
            try:
                self.db_manager.add_violation(
                    self.attendance_id,
                    violation_data['start_time'],
                    violation_data['end_time'],
                    violation_data['description']
                )
                QMessageBox.information(self, "Sukses", "Pelanggaran berhasil ditambahkan")
                self.load_violations()  # Refresh table
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Gagal menambah pelanggaran:\n{str(e)}")
    
    def edit_violation(self):
        """Edit selected violation"""
        selected_rows = self.table.selectionModel().selectedRows()
        if not selected_rows:
            return
        
        row = selected_rows[0].row()
        violation_id = self.table.item(row, 0).data(Qt.UserRole)
        
        # Get current data
        start_time = self.table.item(row, 0).text()
        end_time = self.table.item(row, 1).text()
        description = self.table.item(row, 2).text()
        
        dialog = ViolationEditDialog(start_time, end_time, description, self)
        if dialog.exec() == QDialog.Accepted:
            violation_data = dialog.get_violation_data()
            try:
                self.db_manager.update_violation(
                    violation_id,
                    violation_data['start_time'],
                    violation_data['end_time'],
                    violation_data['description']
                )
                QMessageBox.information(self, "Sukses", "Pelanggaran berhasil diupdate")
                self.load_violations()  # Refresh table
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Gagal mengupdate pelanggaran:\n{str(e)}")
    
    def delete_violation(self):
        """Delete selected violation"""
        selected_rows = self.table.selectionModel().selectedRows()
        if not selected_rows:
            return
        
        row = selected_rows[0].row()
        violation_id = self.table.item(row, 0).data(Qt.UserRole)
        start_time = self.table.item(row, 0).text()
        end_time = self.table.item(row, 1).text()
        description = self.table.item(row, 2).text()
        
        reply = QMessageBox.question(
            self, "Konfirmasi Hapus",
            f"Apakah Anda yakin ingin menghapus pelanggaran ini?\n\n"
            f"Waktu: {start_time} - {end_time}\n"
            f"Keterangan: {description}",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            try:
                self.db_manager.delete_violation(violation_id)
                QMessageBox.information(self, "Sukses", "Pelanggaran berhasil dihapus")
                self.load_violations()  # Refresh table
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Gagal menghapus pelanggaran:\n{str(e)}")

class ViolationEditDialog(QDialog):
    def __init__(self, start_time="", end_time="", description="", parent=None):
        super().__init__(parent)
        self.setWindowTitle("Edit Pelanggaran" if start_time else "Tambah Pelanggaran")
        self.setModal(True)
        self.resize(400, 250)
        
        layout = QFormLayout()
        
        # Waktu mulai dan selesai dengan detik
        self.start_time = QTimeEdit()
        self.start_time.setDisplayFormat("HH:mm:ss")
        if start_time:
            from PySide6.QtCore import QTime
            self.start_time.setTime(QTime.fromString(start_time, "HH:mm:ss"))
        layout.addRow("Jam Mulai:", self.start_time)
        
        self.end_time = QTimeEdit()
        self.end_time.setDisplayFormat("HH:mm:ss")
        if end_time:
            from PySide6.QtCore import QTime
            self.end_time.setTime(QTime.fromString(end_time, "HH:mm:ss"))
        layout.addRow("Jam Selesai:", self.end_time)
        
        # Keterangan
        self.description = QTextEdit()
        self.description.setMaximumHeight(100)
        self.description.setPlainText(description)
        layout.addRow("Keterangan:", self.description)
        
        # Buttons
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addRow(buttons)
        
        self.setLayout(layout)
    
    def get_violation_data(self):
        return {
            'start_time': self.start_time.time().toString("HH:mm:ss"),
            'end_time': self.end_time.time().toString("HH:mm:ss"),
            'description': self.description.toPlainText()
        }

class AttendanceInputTab(QWidget):
    def __init__(self, db_manager, main_window=None):
        super().__init__()
        self.db_manager = db_manager
        self.main_window = main_window
        self.current_data = []
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        
        # Top controls
        controls_layout = QHBoxLayout()
        
        # Date picker dengan format Indonesia
        controls_layout.addWidget(QLabel("Tanggal:"))
        self.date_edit = IndonesianDateEdit()
        self.date_edit.setDate(QDate.currentDate())
        self.date_edit.dateChanged.connect(self.load_attendance_data)
        controls_layout.addWidget(self.date_edit)
        
        controls_layout.addStretch()
        
        # Import button
        self.import_btn = QPushButton("Import Excel")
        self.import_btn.clicked.connect(self.import_excel)
        controls_layout.addWidget(self.import_btn)
        
        # Refresh button
        self.refresh_btn = QPushButton("🔄 Refresh Data")
        self.refresh_btn.setToolTip("Muat ulang data dari database untuk tanggal yang dipilih")
        self.refresh_btn.clicked.connect(self.refresh_data)
        controls_layout.addWidget(self.refresh_btn)
        
        # Save button - ganti dengan Save/Update
        self.save_btn = QPushButton("Save/Update Data")
        self.save_btn.setToolTip("Simpan data baru atau update data yang sudah ada")
        self.save_btn.clicked.connect(self.save_to_database)
        self.save_btn.setEnabled(False)
        controls_layout.addWidget(self.save_btn)
        
        # Database status button
        self.db_status_btn = QPushButton("Check DB Status")
        self.db_status_btn.clicked.connect(self.check_database_status)
        controls_layout.addWidget(self.db_status_btn)
        
        layout.addLayout(controls_layout)
        
        # Table
        self.table = QTableWidget()
        self.table.setColumnCount(10)
        self.table.setHorizontalHeaderLabels([
            "Nama Karyawan", "Shift", "Jam Masuk Kerja", "Jam Keluar Kerja", 
            "Jam Masuk Lembur", "Jam Keluar Lembur", "Jam Anomali", "Keterangan", "Kelola Pelanggaran", "Kelola Izin"
        ])
        
        # Make table editable
        self.table.itemChanged.connect(self.on_item_changed)
        
        # Resize columns - ubah ke Interactive agar pengguna dapat mengubah ukuran kolom
        header = self.table.horizontalHeader()
        
        # Set semua kolom ke Interactive (bisa diubah ukurannya oleh user)
        for i in range(10):  # Semua kolom termasuk Shift, Keterangan, Kelola Pelanggaran dan Kelola Izin
            header.setSectionResizeMode(i, QHeaderView.Interactive)
        
        # Set default width untuk kolom
        self.table.setColumnWidth(0, 160)  # Nama Karyawan
        self.table.setColumnWidth(1, 100)  # Shift
        self.table.setColumnWidth(2, 100)  # Jam Masuk Kerja
        self.table.setColumnWidth(3, 100)  # Jam Keluar Kerja
        self.table.setColumnWidth(4, 120)  # Jam Masuk Lembur
        self.table.setColumnWidth(5, 120)  # Jam Keluar Lembur
        self.table.setColumnWidth(6, 120)  # Jam Anomali
        self.table.setColumnWidth(7, 200)  # Keterangan
        self.table.setColumnWidth(8, 120)  # Kelola Pelanggaran
        self.table.setColumnWidth(9, 120)  # Kelola Izin
        
        # Enable stretching table to fill available space
        self.table.horizontalHeader().setStretchLastSection(True)
        
        # Set row height for better readability
        self.table.verticalHeader().setDefaultSectionSize(45)
        
        layout.addWidget(self.table)
        
        # Bottom controls
        bottom_layout = QHBoxLayout()
        
        self.add_violation_btn = QPushButton("Tambah Pelanggaran")
        self.add_violation_btn.clicked.connect(self.add_violation)
        self.add_violation_btn.setEnabled(False)
        bottom_layout.addWidget(self.add_violation_btn)
        
        self.add_leave_btn = QPushButton("Tambah Izin")
        self.add_leave_btn.setStyleSheet("""
            QPushButton {
                background-color: #28a745;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 8px 15px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #218838;
            }
        """)
        self.add_leave_btn.clicked.connect(self.add_leave)
        self.add_leave_btn.setEnabled(False)
        bottom_layout.addWidget(self.add_leave_btn)
        
        bottom_layout.addStretch()
        
        layout.addLayout(bottom_layout)
        
        self.setLayout(layout)
        
        # Load data for current date
        self.load_attendance_data()
    
    def import_excel(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Pilih File Excel", "", "Excel Files (*.xls *.xlsx)"
        )
        
        if file_path:
            # Show loading cursor
            QApplication.setOverrideCursor(Qt.WaitCursor)
            
            try:
                # Clear any previous state/cache
                import gc
                gc.collect()  # Force garbage collection
                
                # Create fresh processor instance
                processor = ExcelProcessor()
                
                # Process with explicit error handling and debugging
                print(f"🔄 Processing file: {file_path}")
                data = processor.process_excel_log(file_path)
                print(f"📊 Processed data count: {len(data) if data else 0}")
                
                if data:
                    # Clear current data first
                    self.current_data = []
                    self.table.setRowCount(0)
                    
                    # Set new data
                    self.current_data = data
                    self.populate_table(data)
                    self.save_btn.setEnabled(True)
                    self.save_btn.setText("Save Data")  # Ubah teks tombol menjadi Save Data
                    self.add_violation_btn.setEnabled(True)
                    self.add_leave_btn.setEnabled(True)
                    
                    print(f"✅ Import successful: {len(data)} employees")
                    QMessageBox.information(self, "Sukses", f"Berhasil import {len(data)} data karyawan")
                else:
                    print("❌ No data processed from Excel file")
                    QMessageBox.warning(self, "Warning", 
                                      "Tidak ada data yang berhasil diproses dari file Excel.\n\n"
                                      "Kemungkinan penyebab:\n"
                                      "• Format file tidak sesuai dengan yang diharapkan\n"
                                      "• File kosong atau corrupt\n"
                                      "• Struktur data berbeda dari format standar\n\n"
                                      "Pastikan file Excel berisi data absensi dengan format yang benar.")
                    
            except FileNotFoundError as e:
                print(f"❌ File not found: {e}")
                QMessageBox.critical(self, "File Tidak Ditemukan", str(e))
            except Exception as e:
                error_msg = str(e)
                print(f"❌ Import error: {error_msg}")
                
                if "OLE2 inconsistency" in error_msg or "file size" in error_msg:
                    QMessageBox.critical(
                        self, "Error Format Excel", 
                        f"File Excel memiliki format yang tidak standar (umum pada file dari alat presensi lama).\n\n"
                        f"Solusi yang bisa dicoba:\n"
                        f"1. Buka file Excel dan Save As dengan format .xlsx\n"
                        f"2. Gunakan Excel versi terbaru untuk menyimpan file\n"
                        f"3. Export ulang dari alat presensi dengan format yang lebih baru\n\n"
                        f"Catatan: File mungkin masih bisa diproses meskipun ada warning.\n\n"
                        f"Detail error: {error_msg}"
                    )
                elif "DATA KOSONG" in error_msg or "tidak mengandung data" in error_msg:
                    QMessageBox.warning(
                        self, "Data Kosong",
                        f"File Excel berhasil dibaca tapi tidak mengandung data absensi yang dapat diproses.\n\n"
                        f"Kemungkinan penyebab:\n"
                        f"1. Format file berbeda dari yang diharapkan\n"
                        f"2. File kosong atau tidak mengandung data karyawan\n"
                        f"3. Struktur data dalam file berubah\n\n"
                        f"Pastikan file Excel berisi data absensi dengan format yang benar."
                    )
                else:
                    QMessageBox.critical(self, "Error", f"Gagal membaca file Excel:\n{error_msg}")
            finally:
                # Restore normal cursor
                QApplication.restoreOverrideCursor()
                
                # Additional cleanup
                try:
                    import gc
                    gc.collect()  # Clean up any remaining objects
                except:
                    pass
    
    def populate_table(self, data):
        self.table.setRowCount(len(data))
        
        # Get all shifts for dropdown
        shifts = self.db_manager.get_all_shifts()
        
        for row, item in enumerate(data):
            # Nama (read-only)
            name_item = QTableWidgetItem(item['Nama'])
            name_item.setFlags(name_item.flags() & ~Qt.ItemIsEditable)
            self.table.setItem(row, 0, name_item)
            
            # Shift dropdown
            shift_combo = QComboBox()
            shift_combo.setStyleSheet("""
                QComboBox {
                    font-size: 11px;
                    padding: 4px;
                    border: 1px solid #ced4da;
                    border-radius: 3px;
                }
            """)
            for shift in shifts:
                shift_combo.addItem(shift['name'], shift['id'])
            
            # Set current shift - prioritas: shift_id dari data, lalu shift default karyawan, lalu shift 1
            current_shift_id = item.get('shift_id')
            if not current_shift_id:
                # Cari shift default karyawan dari database
                try:
                    employee_data = self.db_manager.get_employee_by_name(item['Nama'])
                    if employee_data:
                        current_shift_id = employee_data.get('shift_id', 1)
                    else:
                        current_shift_id = 1
                except:
                    current_shift_id = 1
            
            shift_index = shift_combo.findData(current_shift_id)
            if shift_index >= 0:
                shift_combo.setCurrentIndex(shift_index)
            
            # Update current_data dengan shift_id yang benar
            item['shift_id'] = current_shift_id
            
            # Connect change event
            shift_combo.currentIndexChanged.connect(
                lambda idx, r=row, combo=shift_combo: self.on_shift_changed(r, combo)
            )
            self.table.setCellWidget(row, 1, shift_combo)
            
            # Jam fields (editable)
            self.table.setItem(row, 2, QTableWidgetItem(item['Jam Masuk'] or ""))
            self.table.setItem(row, 3, QTableWidgetItem(item['Jam Keluar'] or ""))
            self.table.setItem(row, 4, QTableWidgetItem(item['Jam Masuk Lembur'] or ""))
            self.table.setItem(row, 5, QTableWidgetItem(item['Jam Keluar Lembur'] or ""))
            
            # Jam Anomali (read-only, display as comma-separated)
            anomali_text = ", ".join(item['Jam Anomali']) if item['Jam Anomali'] else ""
            anomali_item = QTableWidgetItem(anomali_text)
            anomali_item.setFlags(anomali_item.flags() & ~Qt.ItemIsEditable)
            self.table.setItem(row, 6, anomali_item)
            
            # Keterangan (editable)
            keterangan_text = item.get('keterangan', '') or ""
            keterangan_item = QTableWidgetItem(keterangan_text)
            self.table.setItem(row, 7, keterangan_item)
            
            # Kelola Pelanggaran button dengan info total pelanggaran
            violations_widget = QWidget()
            violations_layout = QHBoxLayout(violations_widget)
            violations_layout.setContentsMargins(2, 2, 2, 2)
            violations_layout.setSpacing(5)
            
            # Tombol Kelola
            manage_btn = QPushButton("Kelola")
            manage_btn.setMaximumWidth(80)
            manage_btn.setStyleSheet("""
                QPushButton {
                    font-size: 11px;
                    padding: 4px 8px;
                    border: 1px solid #ced4da;
                    border-radius: 3px;
                    background-color: #f8f9fa;
                }
                QPushButton:hover {
                    background-color: #e2e6ea;
                }
            """)
            manage_btn.clicked.connect(lambda checked, r=row: self.manage_violations(r))
            violations_layout.addWidget(manage_btn)
            
            # Label total pelanggaran
            violations_count = 0
            if 'id' in item and item['id']:
                try:
                    violations = self.db_manager.get_violations_by_attendance(item['id'])
                    violations_count = len(violations) if violations else 0
                except:
                    pass
            
            count_label = QLabel(f"({violations_count} pelanggaran)")
            if violations_count > 0:
                count_label.setStyleSheet("color: red; font-weight: bold; font-size: 10px;")
            else:
                count_label.setStyleSheet("color: gray; font-size: 10px;")
            violations_layout.addWidget(count_label)
            
            violations_layout.addStretch()
            self.table.setCellWidget(row, 8, violations_widget)
            
            # Kelola Izin button dengan info total izin
            leaves_widget = QWidget()
            leaves_layout = QHBoxLayout(leaves_widget)
            leaves_layout.setContentsMargins(2, 2, 2, 2)
            leaves_layout.setSpacing(5)
            
            # Tombol Kelola
            manage_leave_btn = QPushButton("Kelola")
            manage_leave_btn.setMaximumWidth(80)
            manage_leave_btn.setStyleSheet("""
                QPushButton {
                    background-color: #28a745;
                    color: white;
                    border: none;
                    border-radius: 3px;
                    font-size: 11px;
                    padding: 4px 8px;
                }
                QPushButton:hover {
                    background-color: #218838;
                }
            """)
            manage_leave_btn.clicked.connect(lambda checked, r=row: self.manage_leaves(r))
            leaves_layout.addWidget(manage_leave_btn)
            
            # Label total izin
            leaves_count = 0
            if 'Nama' in item:
                try:
                    # Get employee ID
                    employee_data = self.db_manager.get_employee_by_name(item['Nama'])
                    if employee_data:
                        current_date = self.date_edit.date().toString("yyyy-MM-dd")
                        leaves = self.db_manager.get_leaves_by_employee_date(employee_data['id'], current_date)
                        leaves_count = len(leaves) if leaves else 0
                except:
                    pass
            
            count_leave_label = QLabel(f"({leaves_count} izin)")
            if leaves_count > 0:
                count_leave_label.setStyleSheet("color: green; font-weight: bold; font-size: 10px;")
            else:
                count_leave_label.setStyleSheet("color: gray; font-size: 10px;")
            leaves_layout.addWidget(count_leave_label)
            
            leaves_layout.addStretch()
            self.table.setCellWidget(row, 9, leaves_widget)
            
            # Apply green highlight if employee has leave on this date
            if leaves_count > 0:
                for col in range(10):  # All columns
                    cell_item = self.table.item(row, col)
                    if cell_item:
                        cell_item.setBackground(QColor(200, 255, 200))  # Light green
    
    def on_item_changed(self, item):
        # Update current_data when table is edited
        row = item.row()
        col = item.column()
        
        if row < len(self.current_data):
            field_map = {
                2: 'Jam Masuk',
                3: 'Jam Keluar', 
                4: 'Jam Masuk Lembur',
                5: 'Jam Keluar Lembur',
                7: 'keterangan'  # Kolom keterangan
            }
            
            if col in field_map:
                value = item.text().strip() if item.text().strip() else None
                self.current_data[row][field_map[col]] = value
                
                # Jika keterangan diubah dan data sudah ada di database, update langsung
                if col == 7 and 'id' in self.current_data[row] and self.current_data[row]['id']:
                    try:
                        self.db_manager.update_attendance_keterangan(
                            self.current_data[row]['id'], 
                            value or ''
                        )
                        print(f"✅ Keterangan updated for {self.current_data[row]['Nama']}")
                    except Exception as e:
                        print(f"❌ Failed to update keterangan: {e}")
                        QMessageBox.warning(self, "Warning", f"Gagal update keterangan: {str(e)}")
    
    def on_shift_changed(self, row, combo):
        """Handle shift change for a specific row"""
        if row < len(self.current_data):
            shift_id = combo.currentData()
            shift_name = combo.currentText()
            
            # Update current_data
            self.current_data[row]['shift_id'] = shift_id
            self.current_data[row]['shift_name'] = shift_name
            
            # If this is existing data (has id), update database immediately
            if 'id' in self.current_data[row] and self.current_data[row]['id']:
                try:
                    self.db_manager.update_attendance_shift(self.current_data[row]['id'], shift_id)
                    print(f"✅ Shift updated for {self.current_data[row]['Nama']}: {shift_name}")
                except Exception as e:
                    print(f"❌ Failed to update shift: {e}")
                    QMessageBox.warning(self, "Warning", f"Gagal update shift: {str(e)}")
    
    def save_to_database(self):
        try:
            # Check database status first
            status = check_database_status()
            if status['locked']:
                reply = QMessageBox.question(
                    self, "Database Locked", 
                    "Database sedang terkunci. Apakah ingin mencoba force unlock?",
                    QMessageBox.Yes | QMessageBox.No
                )
                
                if reply == QMessageBox.Yes:
                    if force_unlock_database():
                        QMessageBox.information(self, "Success", "Database berhasil di-unlock!")
                    else:
                        QMessageBox.critical(self, "Error", "Gagal unlock database!")
                        return
                else:
                    return
            
            selected_date = self.date_edit.date().toString("yyyy-MM-dd")
            
            # Check if data already exists for this date
            existing_data = self.db_manager.get_attendance_by_date(selected_date)
            save_mode = 'replace'  # Default mode
            
            if existing_data:
                # Get summary of existing data
                summary = self.db_manager.get_attendance_summary_by_date(selected_date)
                
                # Create custom dialog for save options
                msg = QMessageBox()
                msg.setWindowTitle("Data Sudah Ada")
                msg.setIcon(QMessageBox.Question)
                msg.setText(f"Data absensi untuk tanggal {selected_date} sudah ada!\n\n"
                           f"Data yang ada: {summary['total_employees']} karyawan "
                           f"({summary['hadir']} hadir, {summary['tidak_hadir']} tidak hadir, {summary['lembur']} lembur)\n"
                           f"Data baru: {len(self.current_data)} karyawan\n\n"
                           "Pilih cara penyimpanan:")
                
                # Add custom buttons
                replace_btn = msg.addButton("Timpa Semua", QMessageBox.DestructiveRole)
                merge_btn = msg.addButton("Gabung/Update", QMessageBox.AcceptRole)
                add_only_btn = msg.addButton("Tambah Baru Saja", QMessageBox.AcceptRole)
                cancel_btn = msg.addButton("Batal", QMessageBox.RejectRole)
                
                msg.setDefaultButton(merge_btn)
                msg.exec()
                
                clicked_button = msg.clickedButton()
                
                if clicked_button == replace_btn:
                    save_mode = 'replace'
                elif clicked_button == merge_btn:
                    save_mode = 'merge'
                elif clicked_button == add_only_btn:
                    save_mode = 'insert_only'
                else:  # cancel_btn
                    return
            
            self.db_manager.save_attendance_data(selected_date, self.current_data, save_mode)
            
            # Show appropriate success message
            if existing_data:
                if save_mode == 'replace':
                    action = "ditimpa seluruhnya"
                elif save_mode == 'merge':
                    action = "digabung/diupdate"
                else:  # insert_only
                    action = "ditambahkan (data baru saja)"
            else:
                action = "disimpan"
                
            QMessageBox.information(self, "Sukses", f"Data berhasil {action} ke database")
            self.save_btn.setEnabled(False)
            
            # Refresh report tab
            if self.main_window:
                self.main_window.refresh_report_tab()
            
        except Exception as e:
            error_msg = str(e)
            if "database is locked" in error_msg.lower():
                reply = QMessageBox.question(
                    self, "Database Locked", 
                    f"Database terkunci: {error_msg}\n\nApakah ingin mencoba force unlock dan retry?",
                    QMessageBox.Yes | QMessageBox.No
                )
                
                if reply == QMessageBox.Yes:
                    if force_unlock_database():
                        # Retry save with default mode
                        try:
                            selected_date = self.date_edit.date().toString("yyyy-MM-dd")
                            self.db_manager.save_attendance_data(selected_date, self.current_data, 'replace')
                            QMessageBox.information(self, "Sukses", "Data berhasil disimpan setelah unlock database")
                            self.save_btn.setEnabled(False)
                            
                            # Refresh report tab
                            if self.main_window:
                                self.main_window.refresh_report_tab()
                        except Exception as retry_e:
                            QMessageBox.critical(self, "Error", f"Gagal menyimpan data setelah unlock:\n{str(retry_e)}")
                    else:
                        QMessageBox.critical(self, "Error", "Gagal unlock database!")
            else:
                QMessageBox.critical(self, "Error", f"Gagal menyimpan data:\n{error_msg}")
    
    def load_attendance_data(self):
        # Load existing data from database for selected date
        selected_date = self.date_edit.date().toString("yyyy-MM-dd")
        data = self.db_manager.get_attendance_by_date(selected_date)
        
        if data:
            self.current_data = data
            self.populate_table(data)
            self.add_violation_btn.setEnabled(True)
            self.add_leave_btn.setEnabled(True)
            # Aktifkan tombol Save/Update jika data sudah ada
            self.save_btn.setEnabled(True)
            self.save_btn.setText("Update Data")
        else:
            self.table.setRowCount(0)
            self.current_data = []
            self.add_violation_btn.setEnabled(False)
            self.add_leave_btn.setEnabled(False)
            # Reset tombol Save/Update
            self.save_btn.setEnabled(False)
            self.save_btn.setText("Save/Update Data")
    
    def refresh_data(self):
        """Refresh data dari database untuk tanggal yang dipilih"""
        try:
            # Simpan posisi scroll saat ini
            scroll_pos = self.table.verticalScrollBar().value()
            
            # Ambil tanggal yang dipilih
            selected_date = self.date_edit.date().toString("yyyy-MM-dd")
            
            # Tampilkan loading indicator
            QApplication.setOverrideCursor(Qt.WaitCursor)
            
            # Ambil data terbaru dari database
            data = self.db_manager.get_attendance_by_date(selected_date)
            
            if data:
                self.current_data = data
                self.populate_table(data)
                self.add_violation_btn.setEnabled(True)
                self.add_leave_btn.setEnabled(True)
                self.save_btn.setEnabled(True)
                self.save_btn.setText("Update Data")
                
                # Kembalikan posisi scroll
                self.table.verticalScrollBar().setValue(scroll_pos)
                
                # Tampilkan pesan sukses
                QMessageBox.information(self, "Refresh Berhasil", 
                                       f"Data untuk tanggal {self.date_edit.date().toString('dd MMMM yyyy')} berhasil dimuat ulang.\n\n"
                                       f"Total data: {len(data)} karyawan")
            else:
                self.table.setRowCount(0)
                self.current_data = []
                self.add_violation_btn.setEnabled(False)
                self.add_leave_btn.setEnabled(False)
                self.save_btn.setEnabled(False)
                self.save_btn.setText("Save/Update Data")
                
                # Tampilkan pesan tidak ada data
                QMessageBox.information(self, "Tidak Ada Data", 
                                       f"Tidak ada data absensi untuk tanggal {self.date_edit.date().toString('dd MMMM yyyy')}.")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal memuat data:\n{str(e)}")
        finally:
            # Kembalikan cursor normal
            QApplication.restoreOverrideCursor()
    
    def add_violation(self):
        if not self.current_data:
            QMessageBox.warning(self, "Warning", "Tidak ada data karyawan untuk ditambahkan pelanggaran")
            return
        
        dialog = ViolationDialog(self.current_data, self)
        if dialog.exec() == QDialog.Accepted:
            violation_data = dialog.get_violation_data()
            
            # Find attendance_id based on employee name and date
            selected_date = self.date_edit.date().toString("yyyy-MM-dd")
            attendance_data = self.db_manager.get_attendance_by_date(selected_date)
            
            attendance_id = None
            for att in attendance_data:
                if att['Nama'] == violation_data['employee_name']:
                    attendance_id = att['id']
                    break
            
            if attendance_id:
                try:
                    self.db_manager.add_violation(
                        attendance_id,
                        violation_data['start_time'],
                        violation_data['end_time'], 
                        violation_data['description']
                    )
                    QMessageBox.information(self, "Sukses", "Pelanggaran berhasil ditambahkan")
                except Exception as e:
                    QMessageBox.critical(self, "Error", f"Gagal menambah pelanggaran:\n{str(e)}")
            else:
                QMessageBox.warning(self, "Warning", "Data absensi karyawan tidak ditemukan")
    
    def manage_violations(self, row):
        """Open violation management dialog for selected employee"""
        if row >= len(self.current_data):
            QMessageBox.warning(self, "Warning", "Data karyawan tidak valid")
            return
        
        employee_data = self.current_data[row]
        employee_name = employee_data['Nama']
        
        # Get attendance_id from database
        selected_date = self.date_edit.date().toString("yyyy-MM-dd")
        attendance_data = self.db_manager.get_attendance_by_date(selected_date)
        
        attendance_id = None
        for att in attendance_data:
            if att['Nama'] == employee_name:
                attendance_id = att['id']
                break
        
        if attendance_id:
            dialog = ViolationManagementDialog(self.db_manager, attendance_id, employee_name, self)
            dialog.exec()
            
            # Refresh table to update violation counts
            self.populate_table(self.current_data)
        else:
            QMessageBox.warning(self, "Warning", "Data absensi karyawan tidak ditemukan di database")
    
    def add_leave(self):
        """Add leave for selected employee"""
        if not self.current_data:
            QMessageBox.warning(self, "Warning", "Tidak ada data karyawan untuk ditambahkan izin")
            return
        
        # Get all employees for selection
        employees = []
        for item in self.current_data:
            employees.append(item['Nama'])
        
        dialog = LeaveSelectionDialog(employees, self)
        if dialog.exec() == QDialog.Accepted:
            selected_employee = dialog.get_selected_employee()
            
            # Get employee ID
            try:
                employee_data = self.db_manager.get_employee_by_name(selected_employee)
                if not employee_data:
                    QMessageBox.warning(self, "Warning", f"Data karyawan {selected_employee} tidak ditemukan!")
                    return
                
                current_date = self.date_edit.date().toString("yyyy-MM-dd")
                leave_dialog = LeaveManagementDialog(self.db_manager, employee_data['id'], selected_employee, current_date, self)
                leave_dialog.exec()
                
                # Refresh table to update leave counts and highlights
                self.populate_table(self.current_data)
                
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Gagal membuka dialog izin:\n{str(e)}")
    
    def manage_leaves(self, row):
        """Open leave management dialog for specific row"""
        if row >= len(self.current_data):
            return
        
        employee_name = self.current_data[row]['Nama']
        current_date = self.date_edit.date().toString("yyyy-MM-dd")
        
        # Get employee ID
        try:
            employee_data = self.db_manager.get_employee_by_name(employee_name)
            if not employee_data:
                QMessageBox.warning(self, "Warning", f"Data karyawan {employee_name} tidak ditemukan!")
                return
            
            dialog = LeaveManagementDialog(self.db_manager, employee_data['id'], employee_name, current_date, self)
            dialog.exec()
            
            # Refresh table to update leave counts and highlights
            self.populate_table(self.current_data)
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal membuka dialog izin:\n{str(e)}")
    
    def check_database_status(self):
        """Check dan tampilkan status database"""
        status = check_database_status()
        
        if status['status'] == 'OK':
            msg = f"✅ Database Status: OK\n"
            msg += f"📊 Tables: {status.get('table_count', 0)}\n"
            msg += f"📝 Journal Mode: {status.get('journal_mode', 'unknown')}\n"
            msg += f"🔓 Locked: No"
            QMessageBox.information(self, "Database Status", msg)
            
        elif status['locked']:
            msg = f"🔒 Database Status: LOCKED\n"
            msg += f"❌ Error: {status.get('error', 'Unknown')}\n\n"
            msg += "Apakah ingin mencoba force unlock?"
            
            reply = QMessageBox.question(
                self, "Database Locked", msg,
                QMessageBox.Yes | QMessageBox.No
            )
            
            if reply == QMessageBox.Yes:
                if force_unlock_database():
                    QMessageBox.information(self, "Success", "✅ Database berhasil di-unlock!")
                else:
                    QMessageBox.critical(self, "Error", "❌ Gagal unlock database!")
        else:
            msg = f"❌ Database Status: ERROR\n"
            msg += f"Error: {status.get('error', 'Unknown')}"
            QMessageBox.critical(self, "Database Error", msg)

class LaporanTab(QWidget):
    """Tab Laporan dengan berbagai sub-menu laporan"""
    def __init__(self, db_manager, main_window=None):
        super().__init__()
        self.db_manager = db_manager
        self.main_window = main_window
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        
        # Header
        header_label = QLabel("📊 SISTEM LAPORAN ABSENSI")
        header_label.setAlignment(Qt.AlignCenter)
        header_label.setStyleSheet("""
            QLabel {
                font-size: 24px;
                font-weight: bold;
                color: #2c3e50;
                padding: 20px;
                background-color: #ecf0f1;
                border-radius: 10px;
                margin-bottom: 20px;
            }
        """)
        layout.addWidget(header_label)
        
        # Scroll area untuk menu buttons
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        
        # Widget container untuk buttons
        container = QWidget()
        grid_layout = QGridLayout(container)
        grid_layout.setSpacing(15)
        
        # Menu buttons dengan style yang menarik
        self.create_menu_buttons(grid_layout)
        
        scroll.setWidget(container)
        layout.addWidget(scroll)
        
        self.setLayout(layout)
    
    def create_menu_buttons(self, layout):
        """Membuat menu buttons untuk berbagai jenis laporan"""
        
        # Data menu buttons
        menu_items = [
            {
                'title': '👥 Laporan Masuk\nSemua Karyawan',
                'description': 'Laporan kehadiran semua karyawan\ndalam periode tertentu',
                'color': '#3498db',
                'action': self.open_laporan_masuk_semua
            },
            {
                'title': '⚠️ Laporan Pelanggaran\nSemua Karyawan', 
                'description': 'Laporan pelanggaran dan keterlambatan\nsemua karyawan',
                'color': '#e74c3c',
                'action': self.open_laporan_pelanggaran_semua
            },
            {
                'title': '👤 Laporan Karyawan\nSatuan',
                'description': 'Laporan detail untuk\nkaryawan individual',
                'color': '#27ae60',
                'action': self.open_laporan_karyawan_satuan
            },
            {
                'title': '📈 Laporan Overtime\nSemua Karyawan',
                'description': 'Laporan overtime dan loyalitas\nsemua karyawan',
                'color': '#f39c12',
                'action': self.open_laporan_overtime_semua
            },
            {
                'title': '📅 Laporan Bulanan\nRekap Absensi',
                'description': 'Rekap absensi bulanan\ndengan statistik lengkap',
                'color': '#9b59b6',
                'action': self.open_laporan_bulanan
            },
            {
                'title': '🏆 Laporan Kinerja\nKehadiran',
                'description': 'Analisis kinerja kehadiran\ndan ranking karyawan',
                'color': '#1abc9c',
                'action': self.open_laporan_kinerja
            }
        ]
        
        # Arrange buttons in grid (2 columns)
        row = 0
        col = 0
        for item in menu_items:
            btn = self.create_styled_button(
                item['title'], 
                item['description'], 
                item['color'],
                item['action']
            )
            layout.addWidget(btn, row, col)
            
            col += 1
            if col >= 2:  # 2 columns
                col = 0
                row += 1
    
    def create_styled_button(self, title, description, color, action):
        """Membuat button dengan style yang menarik"""
        btn = QPushButton()
        btn.setFixedSize(300, 120)
        btn.setCursor(Qt.PointingHandCursor)
        
        # Set text
        btn.setText(f"{title}\n\n{description}")
        
        # Set style
        btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {color};
                color: white;
                border: none;
                border-radius: 15px;
                font-size: 12px;
                font-weight: bold;
                text-align: center;
                padding: 10px;
            }}
            QPushButton:hover {{
                background-color: {self.darken_color(color)};
            }}
            QPushButton:pressed {{
                background-color: {self.darken_color(color, 0.8)};
            }}
        """)
        
        # Connect action
        btn.clicked.connect(action)
        
        return btn
    
    def darken_color(self, hex_color, factor=0.8):
        """Menggelapkan warna untuk efek hover"""
        hex_color = hex_color.lstrip('#')
        rgb = tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
        darkened = tuple(int(c * factor) for c in rgb)
        return f"#{darkened[0]:02x}{darkened[1]:02x}{darkened[2]:02x}"
    
    # Action methods untuk setiap menu
    def open_laporan_masuk_semua(self):
        """Buka laporan masuk semua karyawan"""
        dialog = LaporanMasukSemuaDialog(self.db_manager, self)
        dialog.exec()
    
    def open_laporan_pelanggaran_semua(self):
        """Buka laporan pelanggaran semua karyawan"""
        dialog = LaporanPelanggaranSemuaDialog(self.db_manager, self)
        dialog.exec()
    
    def open_laporan_karyawan_satuan(self):
        """Buka laporan karyawan satuan (existing ReportTab)"""
        dialog = LaporanKaryawanSatuanDialog(self.db_manager, self)
        dialog.exec()
    
    def open_laporan_overtime_semua(self):
        """Buka laporan overtime semua karyawan"""
        dialog = LaporanOvertimeSemuaDialog(self.db_manager, self)
        dialog.exec()
    
    def open_laporan_bulanan(self):
        """Buka laporan bulanan"""
        dialog = LaporanBulananDialog(self.db_manager, self)
        dialog.exec()
    
    def open_laporan_kinerja(self):
        """Buka laporan kinerja kehadiran"""
        dialog = LaporanKinerjaDialog(self.db_manager, self)
        dialog.exec()


class ReportTab(QWidget):
    def __init__(self, db_manager):
        super().__init__()
        self.db_manager = db_manager
        self.init_ui()
    
    def init_ui(self):
        # Main splitter
        splitter = QSplitter(Qt.Horizontal)
        
        # Left panel - Settings
        left_panel = self.create_settings_panel()
        splitter.addWidget(left_panel)
        
        # Right panel - Reports
        right_panel = self.create_report_panel()
        splitter.addWidget(right_panel)
        
        # Set splitter proportions
        splitter.setSizes([300, 500])
        
        layout = QVBoxLayout()
        layout.addWidget(splitter)
        self.setLayout(layout)
        
        # Load employees
        self.load_employees()
    
    def create_settings_panel(self):
        group = QGroupBox("Info Shift Karyawan")
        layout = QVBoxLayout()
        
        # Info label
        info_label = QLabel("Shift akan otomatis diambil berdasarkan karyawan yang dipilih di panel laporan.")
        info_label.setWordWrap(True)
        info_label.setStyleSheet("QLabel { color: #666; font-style: italic; }")
        layout.addWidget(info_label)
        
        # Shift info display
        self.shift_info_display = QTextEdit()
        self.shift_info_display.setReadOnly(True)
        self.shift_info_display.setMaximumHeight(400)
        self.shift_info_display.setText("Pilih karyawan di panel laporan untuk melihat info shift")
        layout.addWidget(self.shift_info_display)
        
        group.setLayout(layout)
        return group
    
    def create_report_panel(self):
        group = QGroupBox("Generate Laporan")
        layout = QVBoxLayout()
        
        # Employee selection
        form_layout = QFormLayout()
        
        self.employee_combo = QComboBox()
        self.employee_combo.currentIndexChanged.connect(self.on_employee_changed)
        self.load_employees()
        form_layout.addRow("Pilih Karyawan:", self.employee_combo)
        
        # Date range dengan format Indonesia
        date_layout = QHBoxLayout()
        
        # Tanggal mulai dengan format Indonesia
        self.start_date = IndonesianDateEdit()
        self.start_date.setDate(QDate.currentDate().addDays(-30))
        date_layout.addWidget(self.start_date)
        
        date_layout.addWidget(QLabel(" - "))
        
        # Tanggal akhir dengan format Indonesia
        self.end_date = IndonesianDateEdit()
        self.end_date.setDate(QDate.currentDate())
        date_layout.addWidget(self.end_date)
        
        form_layout.addRow("Periode:", date_layout)
        
        # Buttons layout
        buttons_layout = QHBoxLayout()
        
        # Refresh button
        self.refresh_btn = QPushButton("Refresh Data")
        self.refresh_btn.clicked.connect(self.refresh_employees)
        buttons_layout.addWidget(self.refresh_btn)
        
        # Generate button
        self.generate_btn = QPushButton("Generate Laporan")
        self.generate_btn.clicked.connect(self.generate_report)
        buttons_layout.addWidget(self.generate_btn)
        
        # Export button
        self.export_btn = QPushButton("Export Excel")
        self.export_btn.clicked.connect(self.export_to_excel)
        self.export_btn.setEnabled(False)  # Enable after generate
        buttons_layout.addWidget(self.export_btn)
        
        form_layout.addRow(buttons_layout)
        
        layout.addLayout(form_layout)
        
        # Report table
        self.report_table = QTableWidget()
        self.report_table.setColumnCount(14)  # Tambah kolom Shift dan Loyalitas
        self.report_table.setHorizontalHeaderLabels([
            "Tanggal", "Shift", "Jam Masuk", "Jam Keluar", "Jam Masuk Lembur", "Jam Keluar Lembur",
            "Jam Kerja Total", "Jam Lembur", "Loyalitas", "Overtime", "Keterlambatan", "Status", "Keterangan", "Pelanggaran"
        ])
        
        # Resize columns - ubah ke Interactive agar pengguna dapat mengubah ukuran kolom
        header = self.report_table.horizontalHeader()
        
        # Set semua kolom ke Interactive (bisa diubah ukurannya oleh user)
        for i in range(14):  # Semua kolom termasuk Shift, Loyalitas, keterangan dan pelanggaran
            header.setSectionResizeMode(i, QHeaderView.Interactive)
        
        # Set default width untuk kolom
        self.report_table.setColumnWidth(0, 100)  # Tanggal
        self.report_table.setColumnWidth(1, 80)   # Jam Masuk
        self.report_table.setColumnWidth(2, 80)   # Jam Keluar
        self.report_table.setColumnWidth(3, 120)  # Jam Masuk Lembur
        self.report_table.setColumnWidth(4, 120)  # Jam Keluar Lembur
        self.report_table.setColumnWidth(5, 100)  # Jam Kerja
        self.report_table.setColumnWidth(6, 100)  # Jam Lembur
        self.report_table.setColumnWidth(7, 100)  # Loyalitas
        self.report_table.setColumnWidth(8, 100)  # Overtime
        self.report_table.setColumnWidth(9, 100)  # Keterlambatan
        self.report_table.setColumnWidth(10, 80)  # Status
        self.report_table.setColumnWidth(11, 200) # Keterangan
        self.report_table.setColumnWidth(12, 300) # Pelanggaran
        
        # Enable stretching table to fill available space
        self.report_table.horizontalHeader().setStretchLastSection(True)
        
        # Enable sorting when clicking on headers
        self.report_table.setSortingEnabled(True)
        
        # Enable word wrap untuk semua cells
        self.report_table.setWordWrap(True)
        
        # Set row height yang cukup untuk multiple lines
        self.report_table.verticalHeader().setDefaultSectionSize(45)
        
        layout.addWidget(self.report_table)
        
        # Summary
        self.summary_label = QLabel("Pilih karyawan dan periode untuk melihat laporan")
        self.summary_label.setStyleSheet("font-weight: bold; padding: 10px;")
        layout.addWidget(self.summary_label)
        
        group.setLayout(layout)
        return group
    
    def format_time_duration(self, hours, unit_type="jam"):
        """Format time duration to 'X jam Y menit' or 'X menit' format"""
        if hours == 0:
            return "0 menit"
        
        total_minutes = int(hours * 60)
        jam = total_minutes // 60
        menit = total_minutes % 60
        
        if unit_type == "menit_only":
            return f"{total_minutes} menit"
        
        if jam > 0 and menit > 0:
            return f"{jam} jam {menit} menit"
        elif jam > 0:
            return f"{jam} jam"
        else:
            return f"{menit} menit"
    
    def update_shift_info_display(self, employee_id):
        """Update shift info display based on selected employee - shows per-day shift info"""
        if not employee_id:
            self.shift_info_display.setText("Pilih karyawan untuk melihat info shift")
            return
        
        try:
            employees = self.db_manager.get_employees_with_shifts()
            employee_info = None
            for emp in employees:
                if emp['id'] == employee_id:
                    employee_info = emp
                    break
            
            if not employee_info:
                self.shift_info_display.setText("Karyawan tidak ditemukan!")
                return
            
            # Get attendance data for this employee to show actual shifts used
            start_date = self.start_date.date().toPython().strftime('%Y-%m-%d')
            end_date = self.end_date.date().toPython().strftime('%Y-%m-%d')
            attendance_data = self.db_manager.get_attendance_by_employee_period(employee_id, start_date, end_date)
            
            # Collect unique shifts used in the period
            shifts_used = {}
            for record in attendance_data:
                shift_id = record.get('shift_id', 1)
                if shift_id not in shifts_used:
                    shift_settings = self.db_manager.get_shift_by_id(shift_id)
                    if shift_settings:
                        shifts_used[shift_id] = shift_settings
            
            if not shifts_used:
                # Show default shift if no attendance data
                default_shift = self.db_manager.get_shift_by_id(employee_info.get('shift_id', 1))
                if default_shift:
                    shifts_used[default_shift['id']] = default_shift
            
            # Format shift info
            shift_info = f"""KARYAWAN: {employee_info['name']}

📋 INFORMASI SHIFT PERIODE ({start_date} s/d {end_date}):
⚠️  Shift di-set PER HARI di Input Harian (bukan shift tetap)

SHIFT YANG DIGUNAKAN DALAM PERIODE INI:"""
            
            for shift_id, shift_settings in shifts_used.items():
                shift_info += f"""

🔸 SHIFT: {shift_settings['name']} (ID: {shift_id})
   SENIN - JUMAT:
   • Jam Kerja: {shift_settings['weekday_work_start']} - {shift_settings['weekday_work_end']}
   • Jam Lembur: {shift_settings['weekday_overtime_start']} - {shift_settings['weekday_overtime_end']}
   • Batas Overtime: {shift_settings['weekday_overtime_limit']}
   
   SABTU:
   • Jam Kerja: {shift_settings['saturday_work_start']} - {shift_settings['saturday_work_end']}
   • Jam Lembur: {shift_settings['saturday_overtime_start']} - {shift_settings['saturday_overtime_end']}
   • Batas Overtime: {shift_settings['saturday_overtime_limit']}
   
   PENGATURAN:
   • Toleransi Terlambat: {shift_settings['late_tolerance']} menit
   • Mode Overtime: {shift_settings['overtime_mode'].replace('_', ' ').title()}"""
            
            shift_info += f"""

📝 CATATAN:
• Setiap hari bisa menggunakan shift yang berbeda
• Shift ditentukan saat input data harian
• Laporan ini menggunakan shift sesuai yang di-set per hari
• Minggu: Hitung durasi kerja saja (tidak ada lembur/overtime)"""
            
            self.shift_info_display.setText(shift_info)
            
        except Exception as e:
            self.shift_info_display.setText(f"Error: {str(e)}")
    
    def load_employees(self):
        self.employee_combo.clear()
        employees = self.db_manager.get_all_employees()
        for emp in employees:
            self.employee_combo.addItem(emp['name'], emp['id'])
        
        # Update shift info for first employee if any
        if employees:
            self.update_shift_info_display(employees[0]['id'])
    
    def on_employee_changed(self):
        """Handle employee selection change"""
        employee_id = self.employee_combo.currentData()
        self.update_shift_info_display(employee_id)
    
    def refresh_employees(self):
        """Refresh employee list - dipanggil setelah save data baru"""
        current_selection = self.employee_combo.currentData()
        self.load_employees()
        
        # Restore selection if possible
        if current_selection:
            for i in range(self.employee_combo.count()):
                if self.employee_combo.itemData(i) == current_selection:
                    self.employee_combo.setCurrentIndex(i)
                    break
    
    def generate_report(self):
        if self.employee_combo.count() == 0:
            QMessageBox.warning(self, "Warning", "Tidak ada data karyawan")
            return
        
        employee_id = self.employee_combo.currentData()
        start_date = self.start_date.date().toString("yyyy-MM-dd")
        end_date = self.end_date.date().toString("yyyy-MM-dd")
        
        # Get attendance data
        attendance_data = self.db_manager.get_attendance_by_employee_period(
            employee_id, start_date, end_date
        )
        
        if not attendance_data:
            QMessageBox.information(self, "Info", "Tidak ada data absensi untuk periode yang dipilih")
            return
        
        # Calculate and populate report
        self.calculate_and_populate_report(attendance_data)
        
        # Update shift info display with current period
        self.update_shift_info_display(employee_id)
    
    def calculate_and_populate_report(self, attendance_data):
        # Sekarang menggunakan shift per hari, bukan shift per karyawan
        employee_id = self.employee_combo.currentData()
        employee_name = self.employee_combo.currentText()
        
        self.report_table.setRowCount(len(attendance_data))
        
        total_jam_kerja = 0
        total_jam_lembur = 0
        total_loyalitas = 0
        total_overtime = 0
        total_terlambat = 0
        
        for row, data in enumerate(attendance_data):
            # Populate raw attendance data first
            self.report_table.setItem(row, 0, QTableWidgetItem(data['date']))
            
            # Get shift settings for this specific day (shift per hari)
            shift_id = data.get('shift_id', 1)  # Default to shift 1 if not set
            try:
                shift_settings = self.db_manager.get_shift_by_id(shift_id)
                if not shift_settings:
                    # Fallback to default shift
                    shift_settings = self.db_manager.get_shift_by_id(1)
            except:
                # Fallback to default shift
                shift_settings = self.db_manager.get_shift_by_id(1)
            
            # Add shift name column
            shift_name = shift_settings['name'] if shift_settings else "Default Shift"
            self.report_table.setItem(row, 1, QTableWidgetItem(shift_name))
            
            # Continue with other columns (shifted by 1)
            self.report_table.setItem(row, 2, QTableWidgetItem(data['jam_masuk'] or "-"))
            self.report_table.setItem(row, 3, QTableWidgetItem(data['jam_keluar'] or "-"))
            self.report_table.setItem(row, 4, QTableWidgetItem(data['jam_masuk_lembur'] or "-"))
            self.report_table.setItem(row, 5, QTableWidgetItem(data['jam_keluar_lembur'] or "-"))
            
            # Calculate work hours, overtime, etc. with day detection
            from datetime import datetime
            date_obj = datetime.strptime(data['date'], '%Y-%m-%d')
            day_of_week = date_obj.weekday()  # 0=Monday, 6=Sunday
            
            jam_kerja_normal = self.calculate_work_hours(data, shift_settings, day_of_week)
            jam_lembur = self.calculate_overtime_hours(data, shift_settings, day_of_week)
            loyalitas = self.calculate_loyalitas(data, shift_settings, day_of_week)
            overtime = self.calculate_overtime(data, shift_settings, day_of_week)
            terlambat = self.calculate_lateness(data, shift_settings, day_of_week)
            
            # Calculate total jam kerja = normal + loyalitas + lembur
            jam_kerja_total = self.calculate_total_work_hours(jam_kerja_normal, loyalitas, jam_lembur)
            
            # Populate calculated data with new format "X jam Y menit"
            if day_of_week == 6:  # Sunday - only work duration
                jam_kerja_total_text = self.format_time_duration(jam_kerja_total)
                self.report_table.setItem(row, 6, QTableWidgetItem(jam_kerja_total_text))  # Shifted by 1
                self.report_table.setItem(row, 7, QTableWidgetItem("-"))  # No lembur on Sunday
                self.report_table.setItem(row, 8, QTableWidgetItem("-"))  # No loyalitas on Sunday
                self.report_table.setItem(row, 9, QTableWidgetItem("-"))  # No overtime on Sunday
                self.report_table.setItem(row, 10, QTableWidgetItem("-"))  # No lateness on Sunday
            else:
                # Regular format for other days with new time format
                jam_kerja_total_text = self.format_time_duration(jam_kerja_total)
                jam_lembur_text = self.format_time_duration(jam_lembur)
                loyalitas_text = self.format_time_duration(loyalitas / 60, "menit_only") if loyalitas > 0 else "-"  # loyalitas is in minutes
                overtime_text = self.format_time_duration(overtime, "menit_only")
                terlambat_text = self.format_time_duration(terlambat / 60, "menit_only")  # terlambat is in minutes
                
                self.report_table.setItem(row, 6, QTableWidgetItem(jam_kerja_total_text))  # Total jam kerja
                self.report_table.setItem(row, 7, QTableWidgetItem(jam_lembur_text))
                self.report_table.setItem(row, 8, QTableWidgetItem(loyalitas_text))
                self.report_table.setItem(row, 9, QTableWidgetItem(overtime_text))
                self.report_table.setItem(row, 10, QTableWidgetItem(terlambat_text))
            
            # Check for leaves
            leaves = self.db_manager.get_leaves_by_employee_date(employee_id, data['date'])
            has_leaves = len(leaves) > 0 if leaves else False
            
            # Status
            if has_leaves:
                if len(leaves) > 1:
                    status = f"Izin - Terdapat {len(leaves)} Izin"
                else:
                    status = f"Izin - {leaves[0]['description']}"
                status_item = QTableWidgetItem(status)
                status_item.setBackground(QColor(200, 255, 200))  # Light green background
            else:
                status = "Hadir" if data['jam_masuk'] else "Tidak Hadir"
                status_item = QTableWidgetItem(status)
            
            self.report_table.setItem(row, 11, status_item)  # Shifted by 1
            
            # Keterangan (catatan umum)
            if has_leaves:
                # For leaves, show in keterangan as well
                if len(leaves) > 1:
                    keterangan = f"Izin - Terdapat {len(leaves)} Izin"
                else:
                    keterangan = f"Izin - {leaves[0]['description']}"
            else:
                keterangan = data.get('keterangan', '') or "-"  # Default kosong
            
            # Kolom Pelanggaran (khusus pelanggaran)
            pelanggaran = "-"  # Default kosong
            
            # Get violations for this attendance record
            if 'id' in data and data['id']:
                violations = self.db_manager.get_violations_by_attendance(data['id'])
                if violations:
                    # Format: setiap pelanggaran dalam baris terpisah (newline)
                    # "12:30:00-23:00:00 Tidur\n14:30:00-15:00:00 makan"
                    violation_details = []
                    for violation in violations:
                        start_time = violation['start_time']  # Already in HH:mm:ss format
                        end_time = violation['end_time']      # Already in HH:mm:ss format
                        description = violation['description']
                        violation_details.append(f"{start_time}-{end_time} {description}")
                    
                    # Untuk kolom keterangan tetap kosong atau bisa diisi catatan lain
                    # Untuk kolom pelanggaran diisi dengan detail pelanggaran
                    pelanggaran = "\n".join(violation_details)  # Use newline instead of " | "
            
            # Set keterangan dengan word wrap untuk text panjang
            keterangan_item = QTableWidgetItem(keterangan)
            keterangan_item.setToolTip(keterangan)  # Tooltip untuk text panjang
            self.report_table.setItem(row, 12, keterangan_item)  # Shifted by 1
            
            # Set pelanggaran dengan word wrap untuk text panjang
            pelanggaran_item = QTableWidgetItem(pelanggaran)
            pelanggaran_item.setToolTip(pelanggaran)  # Tooltip untuk text panjang
            if pelanggaran != "-":
                pelanggaran_item.setForeground(QColor(255, 0, 0))  # Warna merah untuk pelanggaran
            self.report_table.setItem(row, 13, pelanggaran_item)  # Shifted by 1
            
            # Add to totals (exclude Sunday from lembur, loyalitas, overtime, lateness)
            total_jam_kerja += jam_kerja_total  # Use total jam kerja
            if day_of_week != 6:  # Not Sunday
                total_jam_lembur += jam_lembur
                total_loyalitas += loyalitas
                total_overtime += overtime
                total_terlambat += terlambat
        
        # Update summary with new format including loyalitas
        employee_name = self.employee_combo.currentText()
        
        # Format totals using the new time format
        total_kerja_text = self.format_time_duration(total_jam_kerja)
        total_lembur_text = self.format_time_duration(total_jam_lembur)
        total_loyalitas_text = self.format_time_duration(total_loyalitas / 60, "menit_only") if total_loyalitas > 0 else "0 menit"  # loyalitas is in minutes
        total_overtime_text = self.format_time_duration(total_overtime, "menit_only")
        total_terlambat_text = self.format_time_duration(total_terlambat / 60, "menit_only")  # terlambat is in minutes
        
        summary_text = (f"Laporan: {employee_name} | "
                       f"Total Kerja: {total_kerja_text} | "
                       f"Total Lembur: {total_lembur_text} | "
                       f"Total Loyalitas: {total_loyalitas_text} | "
                       f"Total Overtime: {total_overtime_text} | "
                       f"Total Terlambat: {total_terlambat_text}")
        
        self.summary_label.setText(summary_text)
        
        # Enable export button after successful report generation
        self.export_btn.setEnabled(True)
    
    def calculate_work_hours(self, data, shift_settings, day_of_week):
        """Calculate work hours based on shift schedule, not actual clock in/out times"""
        if not data['jam_masuk'] or not data['jam_keluar']:
            return 0.0
        
        try:
            jam_masuk_aktual = datetime.strptime(data['jam_masuk'], "%H:%M")
            jam_keluar_aktual = datetime.strptime(data['jam_keluar'], "%H:%M")
            
            # For Sunday (6), just return actual hours worked (no shift schedule)
            if day_of_week == 6:
                if jam_keluar_aktual > jam_masuk_aktual:
                    diff = jam_keluar_aktual - jam_masuk_aktual
                    return diff.total_seconds() / 3600
                return 0.0
            
            # Get shift schedule based on day
            if day_of_week == 5:  # Saturday
                jadwal_masuk = datetime.strptime(shift_settings['saturday_work_start'], "%H:%M")
                jadwal_keluar = datetime.strptime(shift_settings['saturday_work_end'], "%H:%M")
            else:  # Monday to Friday
                jadwal_masuk = datetime.strptime(shift_settings['weekday_work_start'], "%H:%M")
                jadwal_keluar = datetime.strptime(shift_settings['weekday_work_end'], "%H:%M")
            
            # Calculate work hours based on schedule:
            # - Start time: later of (schedule start, actual clock in)
            # - End time: earlier of (schedule end, actual clock out)
            
            jam_mulai_kerja = max(jadwal_masuk, jam_masuk_aktual)
            jam_selesai_kerja = min(jadwal_keluar, jam_keluar_aktual)
            
            # If employee left before schedule end or came after schedule start,
            # calculate based on the overlap with scheduled work hours
            if jam_selesai_kerja > jam_mulai_kerja:
                diff = jam_selesai_kerja - jam_mulai_kerja
                hours = diff.total_seconds() / 3600
                
                # Don't exceed the scheduled work hours
                scheduled_hours = (jadwal_keluar - jadwal_masuk).total_seconds() / 3600
                return min(hours, scheduled_hours)
            
            return 0.0
            
        except Exception as e:
            print(f"Error calculating work hours: {e}")
        return 0.0
    
    def calculate_total_work_hours(self, jam_kerja, loyalitas, jam_lembur):
        """Calculate total work hours = jam kerja normal + loyalitas + lembur"""
        try:
            total = jam_kerja  # Start with normal work hours
            
            # Add loyalitas (convert from minutes to hours)
            if loyalitas > 0:
                total += loyalitas / 60
            
            # Add lembur hours
            if jam_lembur > 0:
                total += jam_lembur
                
            return total
        except:
            return jam_kerja  # Fallback to normal work hours
    
    def calculate_overtime_hours(self, data, shift_settings, day_of_week):
        """Calculate overtime hours based on shift and day"""
        # Sunday has no overtime, only work duration
        if day_of_week == 6:
            return 0.0
            
        if not data['jam_masuk_lembur'] or not data['jam_keluar_lembur']:
            return 0.0
        
        try:
            masuk = datetime.strptime(data['jam_masuk_lembur'], "%H:%M")
            keluar = datetime.strptime(data['jam_keluar_lembur'], "%H:%M")
            
            if keluar > masuk:
                diff = keluar - masuk
                return diff.total_seconds() / 3600
        except:
            pass
        
        return 0.0
    
    def calculate_overtime(self, data, shift_settings, day_of_week):
        """Calculate overtime: jika >1 jam setelah jam kerja maka overtime = 1 jam"""
        # Sunday has no overtime
        if day_of_week == 6:
            return 0.0
            
        if not data['jam_keluar']:
            return 0.0
        
        try:
            keluar = datetime.strptime(data['jam_keluar'], "%H:%M")
            
            # Get scheduled work end time based on day
            if day_of_week == 5:  # Saturday
                jadwal_selesai = datetime.strptime(shift_settings['saturday_work_end'], "%H:%M")
            else:  # Monday-Friday
                jadwal_selesai = datetime.strptime(shift_settings['weekday_work_end'], "%H:%M")
            
            # Calculate overtime with new logic
            if keluar > jadwal_selesai:
                overtime_minutes = (keluar - jadwal_selesai).total_seconds() / 60
                
                # Jika >= 1 jam (60 menit), overtime = 1 jam
                if overtime_minutes >= 60:
                    return 1.0  # 1 jam overtime
                else:
                    # Jika < 1 jam, tidak ada overtime (masuk loyalitas atau tidak ada)
                    return 0.0
                        
        except Exception as e:
            print(f"Error calculating overtime: {e}")
            return 0.0
        
        return 0.0
    
    def calculate_loyalitas(self, data, shift_settings, day_of_week):
        """Calculate loyalitas: 30 menit - 1 jam lebih dari jam kerja normal"""
        # Sunday has no loyalitas calculation
        if day_of_week == 6:
            return 0.0
            
        if not data['jam_keluar']:
            return 0.0
        
        try:
            keluar = datetime.strptime(data['jam_keluar'], "%H:%M")
            
            # Get scheduled work end time based on day
            if day_of_week == 5:  # Saturday
                jadwal_selesai = datetime.strptime(shift_settings['saturday_work_end'], "%H:%M")
            else:  # Monday-Friday
                jadwal_selesai = datetime.strptime(shift_settings['weekday_work_end'], "%H:%M")
            
            # Calculate extra time worked
            if keluar > jadwal_selesai:
                extra_minutes = (keluar - jadwal_selesai).total_seconds() / 60
                    
                # Loyalitas: 30 menit sampai 1 jam (60 menit)
                if 30 <= extra_minutes <= 60:
                    return extra_minutes
                # Jika lebih dari 1 jam, tidak ada loyalitas (sudah masuk overtime)
                elif extra_minutes > 60:
                    return 0.0
                    
        except:
            pass
        
        return 0.0
    
    def calculate_lateness(self, data, shift_settings, day_of_week):
        """Calculate lateness based on shift settings and day"""
        if not data['jam_masuk']:
            return 0.0
        
        try:
            masuk = datetime.strptime(data['jam_masuk'], "%H:%M")
            
            # Get scheduled start time based on day
            if day_of_week == 5:  # Saturday
                jadwal = datetime.strptime(shift_settings['saturday_work_start'], "%H:%M")
            elif day_of_week == 6:  # Sunday - no lateness calculation
                return 0.0
            else:  # Monday-Friday
                jadwal = datetime.strptime(shift_settings['weekday_work_start'], "%H:%M")
            
            # Apply tolerance
            tolerance_minutes = shift_settings['late_tolerance']
            jadwal_with_tolerance = jadwal + timedelta(minutes=tolerance_minutes)
            
            if masuk > jadwal_with_tolerance:
                diff = masuk - jadwal
                return diff.total_seconds() / 60
        except:
            pass
        
        return 0.0
    
    def calculate_loyalitas(self, data, shift_settings, day_of_week):
        """Calculate loyalitas (30 menit - 1 jam lebih dari jam kerja normal)
        Jika >1 jam maka loyalitas = 0 dan overtime = 1 jam"""
        if not data['jam_keluar'] or day_of_week == 6:  # No loyalitas on Sunday
            return 0.0
        
        try:
            keluar = datetime.strptime(data['jam_keluar'], "%H:%M")
            
            # Get scheduled end time based on day
            if day_of_week == 5:  # Saturday
                jadwal_keluar = datetime.strptime(shift_settings['saturday_work_end'], "%H:%M")
            else:  # Monday-Friday
                jadwal_keluar = datetime.strptime(shift_settings['weekday_work_end'], "%H:%M")
            
            if keluar > jadwal_keluar:
                # Calculate extra time in minutes
                extra_minutes = (keluar - jadwal_keluar).total_seconds() / 60
                
                # Loyalitas: 30-60 menit setelah jam pulang normal
                if 30 <= extra_minutes < 60:
                    # Loyalitas sesuai extra time (dalam menit)
                    return extra_minutes
                elif extra_minutes >= 60:
                    # Jika >= 1 jam, loyalitas = 0 (overtime akan jadi 1 jam)
                    return 0.0
                else:
                    # Kurang dari 30 menit, tidak ada loyalitas
                    return 0.0
        except:
            pass
        
        return 0.0
    
    def generate_complete_date_range(self, start_date, end_date, attendance_data):
        """Generate complete date range with empty entries for missing dates"""
        from datetime import datetime, timedelta
        
        # Convert string dates to datetime objects
        start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        end_dt = datetime.strptime(end_date, "%Y-%m-%d")
        
        # Create dictionary of existing attendance data
        attendance_dict = {}
        for data in attendance_data:
            attendance_dict[data['date']] = data
        
        # Generate complete date range
        complete_data = []
        current_date = start_dt
        
        while current_date <= end_dt:
            date_str = current_date.strftime("%Y-%m-%d")
            day_of_week = current_date.weekday()  # 0=Monday, 6=Sunday
            
            if date_str in attendance_dict:
                # Use existing data
                complete_data.append(attendance_dict[date_str])
            else:
                # Create empty entry for missing date
                empty_entry = {
                    'id': None,
                    'date': date_str,
                    'jam_masuk': None,
                    'jam_keluar': None,
                    'jam_masuk_lembur': None,
                    'jam_keluar_lembur': None,
                    'shift_id': 1,  # Default shift
                    'day_name': current_date.strftime("%A"),  # Day name for reference
                    'is_sunday': day_of_week == 6  # Mark Sunday
                }
                complete_data.append(empty_entry)
            
            current_date += timedelta(days=1)
        
        return complete_data
    
    def export_to_excel(self):
        """Export laporan ke file Excel dengan tanggal lengkap termasuk hari kosong"""
        if self.report_table.rowCount() == 0:
            QMessageBox.warning(self, "Warning", "Tidak ada data untuk diekspor. Generate laporan terlebih dahulu.")
            return
        
        # Get file path from user
        employee_name = self.employee_combo.currentText()
        employee_id = self.employee_combo.currentData()
        start_date = self.start_date.date().toString("yyyy-MM-dd")
        end_date = self.end_date.date().toString("yyyy-MM-dd")
        
        # Get attendance data with complete date range
        attendance_data = self.db_manager.get_attendance_by_employee_period(
            employee_id, start_date, end_date
        )
        complete_data = self.generate_complete_date_range(start_date, end_date, attendance_data)
        
        default_filename = f"Laporan_Absensi_{employee_name}_{start_date}_to_{end_date}.xlsx"
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Export Laporan ke Excel", default_filename, "Excel Files (*.xlsx)"
        )
        
        if not file_path:
            return
        
        try:
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Laporan Absensi"
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"), 
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
            
            # Add title
            ws.merge_cells("A1:N1")  # Tambahkan kolom N untuk shift
            title_cell = ws["A1"]
            title_cell.value = f"LAPORAN ABSENSI - {employee_name.upper()}"
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal="center")
            
            # Add period info
            ws.merge_cells("A2:N2")  # Tambahkan kolom N untuk shift
            period_cell = ws["A2"]
            period_cell.value = f"Periode: {start_date} s/d {end_date} (Termasuk hari kosong)"
            period_cell.font = Font(bold=True)
            period_cell.alignment = Alignment(horizontal="center")
            
            # Add headers (tambah kolom Shift)
            headers = [
                "Tanggal", "Shift", "Jam Masuk", "Jam Keluar", "Jam Masuk Lembur", "Jam Keluar Lembur",
                "Jam Kerja", "Jam Lembur", "Loyalitas", "Overtime", "Keterlambatan", "Status", "Keterangan", "Pelanggaran"
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Add data using complete_data (includes empty dates)
            for row, data in enumerate(complete_data):
                # Get shift settings for this day
                shift_id = data.get('shift_id', 1)
                try:
                    shift_settings = self.db_manager.get_shift_by_id(shift_id)
                    if not shift_settings:
                        shift_settings = self.db_manager.get_shift_by_id(1)
                except:
                    shift_settings = self.db_manager.get_shift_by_id(1)
                
                # Calculate values for this day
                from datetime import datetime
                date_obj = datetime.strptime(data['date'], '%Y-%m-%d')
                day_of_week = date_obj.weekday()  # 0=Monday, 6=Sunday
                
                # Basic data (tambah kolom shift)
                excel_row = row + 5
                ws.cell(row=excel_row, column=1).value = data['date']
                
                # Shift name (kolom 2)
                shift_name = shift_settings.get('name', 'Default') if shift_settings else 'Default'
                ws.cell(row=excel_row, column=2).value = shift_name
                
                ws.cell(row=excel_row, column=3).value = data['jam_masuk'] or "-"
                ws.cell(row=excel_row, column=4).value = data['jam_keluar'] or "-"
                ws.cell(row=excel_row, column=5).value = data['jam_masuk_lembur'] or "-"
                ws.cell(row=excel_row, column=6).value = data['jam_keluar_lembur'] or "-"
                
                # Calculate derived values only if there's attendance data
                if data['jam_masuk'] or data['jam_keluar']:
                    jam_kerja = self.calculate_work_hours(data, shift_settings, day_of_week)
                    jam_lembur = self.calculate_overtime_hours(data, shift_settings, day_of_week)
                    loyalitas = self.calculate_loyalitas(data, shift_settings, day_of_week)
                    overtime = self.calculate_overtime(data, shift_settings, day_of_week)
                    terlambat = self.calculate_lateness(data, shift_settings, day_of_week)
                    
                    ws.cell(row=excel_row, column=7).value = self.format_time_duration(jam_kerja)
                    ws.cell(row=excel_row, column=8).value = self.format_time_duration(jam_lembur) if jam_lembur > 0 else "-"
                    ws.cell(row=excel_row, column=9).value = self.format_time_duration(loyalitas / 60, "menit_only") if loyalitas > 0 else "-"
                    ws.cell(row=excel_row, column=10).value = self.format_time_duration(overtime, "menit_only") if overtime > 0 else "-"
                    ws.cell(row=excel_row, column=11).value = self.format_time_duration(terlambat / 60, "menit_only") if terlambat > 0 else "-"
                    ws.cell(row=excel_row, column=12).value = "Hadir"
                    
                    # Highlight orange hanya pada cell keterlambatan (kolom 11)
                    if terlambat > 0:
                        orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                        ws.cell(row=excel_row, column=11).fill = orange_fill  # Hanya kolom keterlambatan
                else:
                    # Empty day
                    ws.cell(row=excel_row, column=7).value = "-"
                    ws.cell(row=excel_row, column=8).value = "-"
                    ws.cell(row=excel_row, column=9).value = "-"
                    ws.cell(row=excel_row, column=10).value = "-"
                    ws.cell(row=excel_row, column=11).value = "-"
                    
                    # Mark Sunday or empty day
                    if day_of_week == 6:  # Sunday
                        ws.cell(row=excel_row, column=12).value = "Minggu"
                        # Highlight Sunday rows with red color
                        red_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                        for col in range(1, 15):  # Tambah 1 kolom untuk shift
                            ws.cell(row=excel_row, column=col).fill = red_fill
                    else:
                        ws.cell(row=excel_row, column=12).value = "Tidak Hadir"
                
                # Keterangan and Pelanggaran
                keterangan_value = data.get('keterangan', '') or "-"
                ws.cell(row=excel_row, column=13).value = keterangan_value
                
                # Get violations if data exists
                pelanggaran = "-"
                if data.get('id'):
                    violations = self.db_manager.get_violations_by_attendance(data['id'])
                    if violations:
                        violation_details = []
                        for violation in violations:
                            violation_details.append(f"{violation['start_time']}-{violation['end_time']} {violation['description']}")
                        pelanggaran = "\n".join(violation_details)
                
                pelanggaran_cell = ws.cell(row=excel_row, column=14)
                pelanggaran_cell.value = pelanggaran
                if pelanggaran != "-":
                    pelanggaran_cell.font = Font(color="FF0000")  # Red color for violations
                
                # Apply borders to all cells
                for col in range(1, 15):  # Tambah 1 kolom untuk shift
                    ws.cell(row=excel_row, column=col).border = border
                    
                    # Center align for certain columns
                    if col in [1, 2, 3, 4, 5, 11]:  # Date, time columns, status
                        ws.cell(row=excel_row, column=col).alignment = Alignment(horizontal="center")
                    elif col == 13:  # Pelanggaran column
                        ws.cell(row=excel_row, column=col).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                        if '\n' in pelanggaran:
                            line_count = pelanggaran.count('\n') + 1
                            ws.row_dimensions[excel_row].height = max(15 * line_count, 30)
                    else:
                        ws.cell(row=excel_row, column=col).alignment = Alignment(horizontal="left", vertical="center")
            
            # Add summary
            summary_row = len(complete_data) + 6
            ws.merge_cells(f"A{summary_row}:N{summary_row}")  # Tambahkan kolom N untuk shift
            summary_cell = ws[f"A{summary_row}"]
            summary_cell.value = f"Laporan lengkap periode {start_date} s/d {end_date} - Total {len(complete_data)} hari (termasuk hari kosong)"
            summary_cell.font = Font(bold=True)
            summary_cell.alignment = Alignment(horizontal="center")
            
            # Add shift rules section
            self.add_shift_rules_to_excel(ws, summary_row + 2)
            
            # Auto-adjust column widths
            from openpyxl.utils import get_column_letter
            for col_num in range(1, 15):  # Columns A to N (1 to 14)
                column_letter = get_column_letter(col_num)
                max_length = 0
                
                # Check all cells in this column
                for row_num in range(1, ws.max_row + 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    # Skip merged cells
                    if hasattr(cell, 'coordinate') and cell.coordinate in ws.merged_cells:
                        continue
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Set column width (minimum 10, maximum 25, except for Pelanggaran column)
                if col_num == 13:  # Kolom Pelanggaran (index 13 dalam Excel)
                    # Kolom pelanggaran dibuat lebih lebar untuk menampung detail pelanggaran
                    adjusted_width = min(max(max_length + 2, 40), 60)
                else:
                    adjusted_width = min(max(max_length + 2, 10), 25)
                
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            wb.save(file_path)
            
            QMessageBox.information(
                self, "Export Berhasil", 
                f"Laporan berhasil diekspor ke:\n{file_path}"
            )
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal mengekspor laporan:\n{str(e)}")
    
    def add_shift_rules_to_excel(self, ws, start_row):
        """Add shift rules section to Excel"""
        try:
            # Get employee shift info
            employee_id = self.employee_combo.currentData()
            employees = self.db_manager.get_employees_with_shifts()
            employee_info = None
            for emp in employees:
                if emp['id'] == employee_id:
                    employee_info = emp
                    break
            
            if not employee_info or not employee_info['shift_id']:
                return
            
            shift_settings = self.db_manager.get_shift_by_id(employee_info['shift_id'])
            if not shift_settings:
                return
            
            # Title
            ws.merge_cells(f"A{start_row}:N{start_row}")  # Tambahkan kolom N untuk shift
            title_cell = ws[f"A{start_row}"]
            title_cell.value = "PERATURAN SHIFT"
            title_cell.font = Font(bold=True, size=12)
            title_cell.alignment = Alignment(horizontal="center")
            
            current_row = start_row + 2
            
            # Shift name
            ws.merge_cells(f"A{current_row}:N{current_row}")  # Tambahkan kolom N untuk shift
            shift_name_cell = ws[f"A{current_row}"]
            shift_name_cell.value = f"SHIFT: {shift_settings['name']}"
            shift_name_cell.font = Font(bold=True)
            shift_name_cell.alignment = Alignment(horizontal="center")
            
            current_row += 2
            
            # Weekday rules
            ws.merge_cells(f"A{current_row}:N{current_row}")  # Tambahkan kolom N untuk shift
            weekday_title = ws[f"A{current_row}"]
            weekday_title.value = "SENIN - JUMAT:"
            weekday_title.font = Font(bold=True)
            
            current_row += 1
            
            weekday_rules = [
                f"• Jam Masuk Kerja: {shift_settings['weekday_work_start']}",
                f"• Jam Keluar Kerja: {shift_settings['weekday_work_end']}",
                f"• Jam Masuk Lembur: {shift_settings['weekday_overtime_start']}",
                f"• Jam Keluar Lembur: {shift_settings['weekday_overtime_end']}",
                f"• Batas Overtime: {shift_settings['weekday_overtime_limit']}"
            ]
            
            for rule in weekday_rules:
                ws.merge_cells(f"A{current_row}:N{current_row}")  # Tambahkan kolom N untuk shift
                rule_cell = ws[f"A{current_row}"]
                rule_cell.value = rule
                current_row += 1
            
            current_row += 1
            
            # Saturday rules
            ws.merge_cells(f"A{current_row}:N{current_row}")  # Tambahkan kolom N untuk shift
            saturday_title = ws[f"A{current_row}"]
            saturday_title.value = "SABTU:"
            saturday_title.font = Font(bold=True)
            
            current_row += 1
            
            saturday_rules = [
                f"• Jam Masuk Kerja: {shift_settings['saturday_work_start']}",
                f"• Jam Keluar Kerja: {shift_settings['saturday_work_end']}",
                f"• Jam Masuk Lembur: {shift_settings['saturday_overtime_start']}",
                f"• Jam Keluar Lembur: {shift_settings['saturday_overtime_end']}",
                f"• Batas Overtime: {shift_settings['saturday_overtime_limit']}"
            ]
            
            for rule in saturday_rules:
                ws.merge_cells(f"A{current_row}:N{current_row}")  # Tambahkan kolom N untuk shift
                rule_cell = ws[f"A{current_row}"]
                rule_cell.value = rule
                current_row += 1
            
            current_row += 1
            
            # Sunday and general rules
            ws.merge_cells(f"A{current_row}:N{current_row}")  # Tambahkan kolom N untuk shift
            sunday_title = ws[f"A{current_row}"]
            sunday_title.value = "MINGGU:"
            sunday_title.font = Font(bold=True)
            
            current_row += 1
            
            ws.merge_cells(f"A{current_row}:N{current_row}")  # Tambahkan kolom N untuk shift
            sunday_rule = ws[f"A{current_row}"]
            sunday_rule.value = "• Hitung durasi kerja saja (tidak ada lembur/overtime)"
            
            current_row += 2
            
            # General settings
            ws.merge_cells(f"A{current_row}:N{current_row}")  # Tambahkan kolom N untuk shift
            general_title = ws[f"A{current_row}"]
            general_title.value = "PENGATURAN UMUM:"
            general_title.font = Font(bold=True)
            
            current_row += 1
            
            general_rules = [
                f"• Toleransi Keterlambatan: {shift_settings['late_tolerance']} menit",
                f"• Mode Overtime: {shift_settings['overtime_mode'].replace('_', ' ').title()}"
            ]
            
            for rule in general_rules:
                ws.merge_cells(f"A{current_row}:N{current_row}")  # Tambahkan kolom N untuk shift
                rule_cell = ws[f"A{current_row}"]
                rule_cell.value = rule
                current_row += 1
                
        except Exception as e:
            print(f"Error adding shift rules: {e}")

class ShiftManagementTab(QWidget):
    def __init__(self, db_manager):
        super().__init__()
        self.db_manager = db_manager
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        
        # Main panel: Shift Settings (CRUD only)
        main_panel = QGroupBox("Pengaturan Shift")
        main_layout = QVBoxLayout(main_panel)
        
        # Info text
        info_label = QLabel("💡 <b>Catatan:</b> Shift diatur per hari di <i>Input Absensi Harian</i>. Di sini hanya untuk mengelola jenis-jenis shift yang tersedia.")
        info_label.setWordWrap(True)
        info_label.setStyleSheet("QLabel { color: #666; background-color: #f0f0f0; padding: 8px; border-radius: 4px; }")
        main_layout.addWidget(info_label)
        
        # Shift selector and CRUD buttons
        shift_selector_layout = QHBoxLayout()
        shift_selector_layout.addWidget(QLabel("Pilih Shift:"))
        self.shift_combo = QComboBox()
        self.shift_combo.currentIndexChanged.connect(self.load_shift_settings)
        shift_selector_layout.addWidget(self.shift_combo)
        
        main_layout.addLayout(shift_selector_layout)
        
        # CRUD buttons
        crud_layout = QHBoxLayout()
        
        create_shift_btn = QPushButton("➕ Buat Shift Baru")
        create_shift_btn.clicked.connect(self.create_shift)
        crud_layout.addWidget(create_shift_btn)
        
        edit_shift_btn = QPushButton("✏️ Edit Shift")
        edit_shift_btn.clicked.connect(self.edit_shift)
        crud_layout.addWidget(edit_shift_btn)
        
        delete_shift_btn = QPushButton("🗑️ Hapus Shift")
        delete_shift_btn.clicked.connect(self.delete_shift)
        crud_layout.addWidget(delete_shift_btn)
        
        main_layout.addLayout(crud_layout)
        
        # Shift details display
        self.shift_details = QTextEdit()
        self.shift_details.setReadOnly(True)
        self.shift_details.setMaximumHeight(400)
        main_layout.addWidget(self.shift_details)
        
        # Add main panel to layout
        layout.addWidget(main_panel)
        
        self.setLayout(layout)
        
        # Load initial data
        self.load_shifts()
    
    # Removed load_data - no longer needed since we only manage shifts, not employee assignments
    
    def load_shifts(self):
        """Load all shifts to combo box"""
        self.shift_combo.clear()
        try:
            shifts = self.db_manager.get_all_shifts()
            for shift in shifts:
                self.shift_combo.addItem(shift['name'], shift['id'])
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal memuat data shift: {str(e)}")
    
    def load_shift_settings(self):
        """Load and display shift settings"""
        shift_id = self.shift_combo.currentData()
        if not shift_id:
            return
        
        try:
            shift = self.db_manager.get_shift_by_id(shift_id)
            if shift:
                details = f"""
SHIFT: {shift['name']}

SENIN - JUMAT:
• Jam Kerja: {shift['weekday_work_start']} - {shift['weekday_work_end']}
• Jam Lembur: {shift['weekday_overtime_start']} - {shift['weekday_overtime_end']}
• Batas Overtime: {shift['weekday_overtime_limit']}

SABTU:
• Jam Kerja: {shift['saturday_work_start']} - {shift['saturday_work_end']}
• Jam Lembur: {shift['saturday_overtime_start']} - {shift['saturday_overtime_end']}
• Batas Overtime: {shift['saturday_overtime_limit']}

PENGATURAN:
• Toleransi Terlambat: {shift['late_tolerance']} menit
• Mode Overtime: {shift['overtime_mode']}
                """.strip()
                
                self.shift_details.setText(details)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal memuat pengaturan shift: {str(e)}")
    
    # Removed load_employees - shift assignment now handled per day in attendance input
    
    # Removed change_employee_shift - shift assignment now handled per day in attendance input
    
    def create_shift(self):
        """Buat shift baru"""
        # Default data untuk shift baru
        default_data = {
            'name': 'Shift Baru',
            'weekday_work_start': '08:00',
            'weekday_work_end': '16:00',
            'weekday_overtime_start': '16:00',
            'weekday_overtime_end': '20:00',
            'weekday_overtime_limit': '17:00',
            'saturday_work_start': '08:00',
            'saturday_work_end': '12:00',
            'saturday_overtime_start': '12:00',
            'saturday_overtime_end': '16:00',
            'saturday_overtime_limit': '13:00',
            'late_tolerance': 15,
            'overtime_mode': 'per_jam'
        }
        
        dialog = ShiftEditDialog(default_data, self, is_create=True)
        if dialog.exec() == QDialog.Accepted:
            try:
                shift_data = dialog.get_shift_data()
                shift_id = self.db_manager.create_shift(shift_data)
                
                QMessageBox.information(self, "Success", f"Shift '{shift_data['name']}' berhasil dibuat!")
                self.load_shifts()
                
                # Select the newly created shift
                for i in range(self.shift_combo.count()):
                    if self.shift_combo.itemData(i) == shift_id:
                        self.shift_combo.setCurrentIndex(i)
                        break
                        
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Gagal membuat shift: {str(e)}")

    def edit_shift(self):
        """Open shift edit dialog"""
        shift_id = self.shift_combo.currentData()
        if not shift_id:
            QMessageBox.warning(self, "Warning", "Pilih shift terlebih dahulu!")
            return
        
        try:
            shift = self.db_manager.get_shift_by_id(shift_id)
            if shift:
                dialog = ShiftEditDialog(shift, self, is_create=False)
                if dialog.exec() == QDialog.Accepted:
                    updated_shift = dialog.get_shift_data()
                    self.db_manager.update_shift(shift_id, updated_shift)
                    QMessageBox.information(self, "Sukses", "Pengaturan shift berhasil diupdate!")
                    self.load_shifts()  # Refresh combo
                    self.load_shift_settings()  # Refresh display
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal mengedit shift: {str(e)}")
    
    def delete_shift(self):
        """Hapus shift yang dipilih"""
        current_shift_id = self.shift_combo.currentData()
        if not current_shift_id:
            QMessageBox.warning(self, "Warning", "Pilih shift terlebih dahulu!")
            return
        
        shift_name = self.shift_combo.currentText()
        
        reply = QMessageBox.question(
            self, 
            "Konfirmasi Hapus", 
            f"Apakah Anda yakin ingin menghapus shift '{shift_name}'?\n\nShift yang sedang digunakan karyawan tidak dapat dihapus.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            try:
                self.db_manager.delete_shift(current_shift_id)
                QMessageBox.information(self, "Success", f"Shift '{shift_name}' berhasil dihapus!")
                self.load_shifts()
                
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Gagal menghapus shift: {str(e)}")


class ShiftEditDialog(QDialog):
    def __init__(self, shift_data, parent=None, is_create=False):
        super().__init__(parent)
        self.shift_data = shift_data
        self.is_create = is_create
        
        if is_create:
            self.setWindowTitle("Buat Shift Baru")
        else:
            self.setWindowTitle(f"Edit {shift_data['name']}")
            
        self.setModal(True)
        self.resize(500, 600)
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        
        # Shift name
        name_layout = QHBoxLayout()
        name_layout.addWidget(QLabel("Nama Shift:"))
        self.name_edit = QLineEdit(self.shift_data['name'])
        name_layout.addWidget(self.name_edit)
        layout.addLayout(name_layout)
        
        # Weekday settings
        weekday_group = QGroupBox("Pengaturan Senin - Jumat")
        weekday_layout = QFormLayout()
        
        self.weekday_work_start = QTimeEdit()
        self.weekday_work_start.setDisplayFormat("HH:mm")
        self.weekday_work_start.setTime(QTime.fromString(self.shift_data['weekday_work_start'], "HH:mm"))
        weekday_layout.addRow("Jam Masuk Kerja:", self.weekday_work_start)
        
        self.weekday_work_end = QTimeEdit()
        self.weekday_work_end.setDisplayFormat("HH:mm")
        self.weekday_work_end.setTime(QTime.fromString(self.shift_data['weekday_work_end'], "HH:mm"))
        weekday_layout.addRow("Jam Keluar Kerja:", self.weekday_work_end)
        
        self.weekday_overtime_start = QTimeEdit()
        self.weekday_overtime_start.setDisplayFormat("HH:mm")
        self.weekday_overtime_start.setTime(QTime.fromString(self.shift_data['weekday_overtime_start'], "HH:mm"))
        weekday_layout.addRow("Jam Masuk Lembur:", self.weekday_overtime_start)
        
        self.weekday_overtime_end = QTimeEdit()
        self.weekday_overtime_end.setDisplayFormat("HH:mm")
        self.weekday_overtime_end.setTime(QTime.fromString(self.shift_data['weekday_overtime_end'], "HH:mm"))
        weekday_layout.addRow("Jam Keluar Lembur:", self.weekday_overtime_end)
        
        self.weekday_overtime_limit = QTimeEdit()
        self.weekday_overtime_limit.setDisplayFormat("HH:mm")
        self.weekday_overtime_limit.setTime(QTime.fromString(self.shift_data['weekday_overtime_limit'], "HH:mm"))
        weekday_layout.addRow("Batas Overtime:", self.weekday_overtime_limit)
        
        weekday_group.setLayout(weekday_layout)
        layout.addWidget(weekday_group)
        
        # Saturday settings
        saturday_group = QGroupBox("Pengaturan Sabtu")
        saturday_layout = QFormLayout()
        
        self.saturday_work_start = QTimeEdit()
        self.saturday_work_start.setDisplayFormat("HH:mm")
        self.saturday_work_start.setTime(QTime.fromString(self.shift_data['saturday_work_start'], "HH:mm"))
        saturday_layout.addRow("Jam Masuk Kerja:", self.saturday_work_start)
        
        self.saturday_work_end = QTimeEdit()
        self.saturday_work_end.setDisplayFormat("HH:mm")
        self.saturday_work_end.setTime(QTime.fromString(self.shift_data['saturday_work_end'], "HH:mm"))
        saturday_layout.addRow("Jam Keluar Kerja:", self.saturday_work_end)
        
        self.saturday_overtime_start = QTimeEdit()
        self.saturday_overtime_start.setDisplayFormat("HH:mm")
        self.saturday_overtime_start.setTime(QTime.fromString(self.shift_data['saturday_overtime_start'], "HH:mm"))
        saturday_layout.addRow("Jam Masuk Lembur:", self.saturday_overtime_start)
        
        self.saturday_overtime_end = QTimeEdit()
        self.saturday_overtime_end.setDisplayFormat("HH:mm")
        self.saturday_overtime_end.setTime(QTime.fromString(self.shift_data['saturday_overtime_end'], "HH:mm"))
        saturday_layout.addRow("Jam Keluar Lembur:", self.saturday_overtime_end)
        
        self.saturday_overtime_limit = QTimeEdit()
        self.saturday_overtime_limit.setDisplayFormat("HH:mm")
        self.saturday_overtime_limit.setTime(QTime.fromString(self.shift_data['saturday_overtime_limit'], "HH:mm"))
        saturday_layout.addRow("Batas Overtime:", self.saturday_overtime_limit)
        
        saturday_group.setLayout(saturday_layout)
        layout.addWidget(saturday_group)
        
        # General settings
        general_group = QGroupBox("Pengaturan Umum")
        general_layout = QFormLayout()
        
        self.late_tolerance = QSpinBox()
        self.late_tolerance.setRange(0, 120)
        self.late_tolerance.setValue(self.shift_data['late_tolerance'])
        general_layout.addRow("Toleransi Terlambat (menit):", self.late_tolerance)
        
        overtime_group = QGroupBox("Mode Overtime")
        overtime_layout = QVBoxLayout()
        
        self.overtime_per_menit = QRadioButton("Per Menit")
        self.overtime_per_jam = QRadioButton("Per Jam (≥60 menit)")
        
        if self.shift_data['overtime_mode'] == 'per_menit':
            self.overtime_per_menit.setChecked(True)
        else:
            self.overtime_per_jam.setChecked(True)
        
        overtime_layout.addWidget(self.overtime_per_menit)
        overtime_layout.addWidget(self.overtime_per_jam)
        overtime_group.setLayout(overtime_layout)
        
        general_layout.addRow(overtime_group)
        general_group.setLayout(general_layout)
        layout.addWidget(general_group)
        
        # Buttons
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
        
        self.setLayout(layout)
    
    def get_shift_data(self):
        return {
            'name': self.name_edit.text(),
            'weekday_work_start': self.weekday_work_start.time().toString("HH:mm"),
            'weekday_work_end': self.weekday_work_end.time().toString("HH:mm"),
            'weekday_overtime_start': self.weekday_overtime_start.time().toString("HH:mm"),
            'weekday_overtime_end': self.weekday_overtime_end.time().toString("HH:mm"),
            'weekday_overtime_limit': self.weekday_overtime_limit.time().toString("HH:mm"),
            'saturday_work_start': self.saturday_work_start.time().toString("HH:mm"),
            'saturday_work_end': self.saturday_work_end.time().toString("HH:mm"),
            'saturday_overtime_start': self.saturday_overtime_start.time().toString("HH:mm"),
            'saturday_overtime_end': self.saturday_overtime_end.time().toString("HH:mm"),
            'saturday_overtime_limit': self.saturday_overtime_limit.time().toString("HH:mm"),
            'late_tolerance': self.late_tolerance.value(),
            'overtime_mode': 'per_menit' if self.overtime_per_menit.isChecked() else 'per_jam'
        }


class LaporanKaryawanSatuanDialog(QDialog):
    """Dialog untuk laporan karyawan satuan menggunakan ReportTab yang sudah ada"""
    def __init__(self, db_manager, parent=None):
        super().__init__(parent)
        self.db_manager = db_manager
        self.setWindowTitle("📊 Laporan Karyawan Satuan")
        self.setModal(True)
        self.resize(1000, 700)
        
        layout = QVBoxLayout()
        
        # Embed existing ReportTab
        self.report_tab = ReportTab(db_manager)
        layout.addWidget(self.report_tab)
        
        # Close button
        close_btn = QPushButton("Tutup")
        close_btn.clicked.connect(self.close)
        close_layout = QHBoxLayout()
        close_layout.addStretch()
        close_layout.addWidget(close_btn)
        layout.addLayout(close_layout)
        
        self.setLayout(layout)


class LaporanMasukSemuaDialog(QDialog):
    """Dialog untuk laporan masuk semua karyawan"""
    def __init__(self, db_manager, parent=None):
        super().__init__(parent)
        self.db_manager = db_manager
        self.setWindowTitle("👥 Laporan Masuk Semua Karyawan")
        self.setModal(True)
        self.resize(1400, 900)
        
        # Data storage
        self.attendance_data = {}
        self.employees = []
        self.date_range = []
        
        layout = QVBoxLayout()
        
        # Header
        header = QLabel("📊 LAPORAN KEHADIRAN SEMUA KARYAWAN")
        header.setAlignment(Qt.AlignCenter)
        header.setStyleSheet("""
            QLabel {
                font-size: 20px;
                font-weight: bold;
                padding: 15px;
                background-color: #f8f9fa;
                border: 2px solid #dee2e6;
                border-radius: 8px;
                color: #495057;
            }
        """)
        layout.addWidget(header)
        
        # Controls
        controls_layout = QHBoxLayout()
        
        controls_layout.addWidget(QLabel("Dari Tanggal:"))
        self.start_date = IndonesianDateEdit()
        self.start_date.setDate(QDate.currentDate().addDays(-30))
        controls_layout.addWidget(self.start_date)
        
        controls_layout.addWidget(QLabel("Sampai Tanggal:"))
        self.end_date = IndonesianDateEdit()
        self.end_date.setDate(QDate.currentDate())
        controls_layout.addWidget(self.end_date)
        
        generate_btn = QPushButton("🔄 Generate Laporan")
        generate_btn.setStyleSheet("""
            QPushButton {
                background-color: #007bff;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px 20px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #0056b3;
            }
        """)
        generate_btn.clicked.connect(self.generate_report)
        controls_layout.addWidget(generate_btn)
        
        self.export_btn = QPushButton("📊 Export Excel")
        self.export_btn.setStyleSheet("""
            QPushButton {
                background-color: #28a745;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px 20px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #1e7e34;
            }
        """)
        self.export_btn.clicked.connect(self.export_excel)
        self.export_btn.setEnabled(False)
        controls_layout.addWidget(self.export_btn)
        
        controls_layout.addStretch()
        layout.addLayout(controls_layout)
        
        # Progress bar (hidden by default)
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        layout.addWidget(self.progress_bar)
        
        # Table with scroll area
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        
        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.horizontalHeader().setStretchLastSection(False)
        self.table.setStyleSheet("""
            QTableWidget {
                gridline-color: #dee2e6;
                background-color: white;
            }
            QTableWidget::item {
                padding: 8px;
                text-align: center;
            }
            QHeaderView::section {
                background-color: #f8f9fa;
                padding: 8px;
                border: 1px solid #dee2e6;
                font-weight: bold;
            }
        """)
        
        scroll_area.setWidget(self.table)
        layout.addWidget(scroll_area)
        
        # Legend
        legend_layout = QHBoxLayout()
        legend_label = QLabel("KETERANGAN: ✅ = Hadir lengkap  |  ⚠️ (orange) = Hadir tidak lengkap  |  📧 (hijau) = Karyawan izin  |  (kosong) = Tidak hadir  |  (merah) = Hari Minggu")
        legend_label.setStyleSheet("""
            QLabel {
                font-size: 12px;
                padding: 8px;
                background-color: #e9ecef;
                border-radius: 5px;
                color: #495057;
            }
        """)
        legend_layout.addWidget(legend_label)
        layout.addLayout(legend_layout)
        
        # Close button
        close_btn = QPushButton("Tutup")
        close_btn.clicked.connect(self.close)
        close_layout = QHBoxLayout()
        close_layout.addStretch()
        close_layout.addWidget(close_btn)
        layout.addLayout(close_layout)
        
        self.setLayout(layout)
    
    def validate_date_range(self):
        """Validasi range tanggal"""
        start = self.start_date.date().toPython()
        end = self.end_date.date().toPython()
        
        if start > end:
            QMessageBox.warning(self, "Error", "Tanggal mulai tidak boleh lebih besar dari tanggal akhir!")
            return False
        
        # Check maksimal 3 bulan
        max_days = 90
        days_diff = (end - start).days + 1
        
        if days_diff > max_days:
            QMessageBox.warning(self, "Error", f"Range tanggal maksimal {max_days} hari!\nRange yang dipilih: {days_diff} hari")
            return False
        
        return True
    
    def generate_report(self):
        """Generate laporan masuk semua karyawan"""
        if not self.validate_date_range():
            return
        
        # Show loading
        self.progress_bar.setVisible(True)
        self.progress_bar.setRange(0, 0)  # Indeterminate progress
        QApplication.processEvents()
        
        try:
            # Get date range
            start_date = self.start_date.date().toPython()
            end_date = self.end_date.date().toPython()
            
            # Generate date range
            self.date_range = []
            current_date = start_date
            while current_date <= end_date:
                self.date_range.append(current_date)
                current_date += timedelta(days=1)
            
            # Get all employees (sorted alphabetically)
            self.employees = self.db_manager.get_all_employees()
            self.employees.sort(key=lambda x: x['name'])
            
            # Get attendance data for all employees in date range
            self.attendance_data = {}
            for employee in self.employees:
                emp_data = self.db_manager.get_attendance_by_employee_period(
                    employee['id'], 
                    start_date.strftime('%Y-%m-%d'), 
                    end_date.strftime('%Y-%m-%d')
                )
                
                # Convert to dict by date for easy lookup
                self.attendance_data[employee['id']] = {}
                for record in emp_data:
                    self.attendance_data[employee['id']][record['date']] = record
            
            # Populate table
            self.populate_attendance_matrix()
            
            # Enable export button
            self.export_btn.setEnabled(True)
            
            QMessageBox.information(self, "Success", f"Laporan berhasil dibuat!\nPeriode: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}\nTotal: {len(self.employees)} karyawan, {len(self.date_range)} hari")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal membuat laporan: {str(e)}")
        finally:
            # Hide loading
            self.progress_bar.setVisible(False)
    
    def populate_attendance_matrix(self):
        """Populate tabel matrix kehadiran"""
        if not self.employees or not self.date_range:
            return
        
        # Setup table dimensions
        # Columns: Nama + Tanggal + Summary
        # Rows: Employees + Summary row
        num_cols = 1 + len(self.date_range) + 1  # Nama + dates + summary
        num_rows = len(self.employees) + 1  # employees + summary row
        
        self.table.setRowCount(num_rows)
        self.table.setColumnCount(num_cols)
        
        # Setup headers
        headers = ["Nama Karyawan"]
        
        # Add date headers with day names
        for date in self.date_range:
            day_names = ["Sen", "Sel", "Rab", "Kam", "Jum", "Sab", "Min"]
            day_name = day_names[date.weekday()]
            date_str = f"{day_name}, {date.strftime('%d/%m')}"
            headers.append(date_str)
        
        headers.append("Total Hadir")
        self.table.setHorizontalHeaderLabels(headers)
        
        # Populate employee rows
        for row, employee in enumerate(self.employees):
            # Employee name
            name_item = QTableWidgetItem(employee['name'])
            name_item.setFlags(name_item.flags() & ~Qt.ItemIsEditable)
            self.table.setItem(row, 0, name_item)
            
            # Attendance data for each date
            total_present = 0
            for col, date in enumerate(self.date_range, 1):
                date_str = date.strftime('%Y-%m-%d')
                attendance = self.attendance_data[employee['id']].get(date_str)
                
                item = QTableWidgetItem()
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                item.setTextAlignment(Qt.AlignCenter)
                
                # Check for leaves first
                leaves = self.db_manager.get_leaves_by_employee_date(employee['id'], date_str)
                has_leaves = len(leaves) > 0 if leaves else False
                
                if has_leaves:
                    # Employee has leave - green background with envelope symbol
                    if len(leaves) > 1:
                        item.setText(f"📧 ({len(leaves)})")
                    else:
                        item.setText("📧")
                    item.setBackground(QColor(200, 255, 200))  # Light green background
                    total_present += 1  # Count leave as present
                elif attendance:
                    # Check if data is complete
                    has_masuk = attendance.get('jam_masuk') and attendance['jam_masuk'].strip()
                    has_keluar = attendance.get('jam_keluar') and attendance['jam_keluar'].strip()
                    
                    if has_masuk and has_keluar:
                        # Complete data - checkmark
                        item.setText("✅")
                        # Check if Sunday for background color
                        if date.weekday() == 6:  # Sunday
                            item.setBackground(QColor(255, 200, 200))  # Light red for Sunday
                        else:
                            item.setBackground(QColor(255, 255, 255))  # White background
                    elif has_masuk or has_keluar:
                        # Incomplete data - error symbol with orange background
                        item.setText("⚠️")
                        if date.weekday() == 6:  # Sunday
                            item.setBackground(QColor(255, 150, 150))  # Darker red for Sunday + incomplete
                        else:
                            item.setBackground(QColor(255, 140, 0))  # Darker orange background for better visibility
                    else:
                        # No attendance data
                        item.setText("")
                        if date.weekday() == 6:  # Sunday
                            item.setBackground(QColor(255, 200, 200))  # Light red for Sunday
                        else:
                            item.setBackground(QColor(255, 255, 255))
                    
                    # Count as present if any data exists
                    if has_masuk or has_keluar:
                        total_present += 1
                else:
                    # No attendance record and no leave
                    item.setText("")
                    # Highlight Sundays in red
                    if date.weekday() == 6:  # Sunday
                        item.setBackground(QColor(255, 200, 200))  # Light red
                    else:
                        item.setBackground(QColor(255, 255, 255))
                
                self.table.setItem(row, col, item)
            
            # Total column
            total_item = QTableWidgetItem(str(total_present))
            total_item.setFlags(total_item.flags() & ~Qt.ItemIsEditable)
            total_item.setTextAlignment(Qt.AlignCenter)
            total_item.setBackground(QColor(240, 248, 255))  # Light blue
            self.table.setItem(row, len(self.date_range) + 1, total_item)
        
        # Summary row (total employees present per date)
        summary_row = len(self.employees)
        
        # Summary row label
        summary_label = QTableWidgetItem("TOTAL HADIR")
        summary_label.setFlags(summary_label.flags() & ~Qt.ItemIsEditable)
        summary_label.setBackground(QColor(240, 248, 255))
        summary_label.setFont(QFont("", 0, QFont.Bold))
        self.table.setItem(summary_row, 0, summary_label)
        
        # Calculate totals per date
        for col, date in enumerate(self.date_range, 1):
            date_str = date.strftime('%Y-%m-%d')
            total_present_on_date = 0
            
            for employee in self.employees:
                attendance = self.attendance_data[employee['id']].get(date_str)
                if attendance:
                    has_masuk = attendance.get('jam_masuk') and attendance['jam_masuk'].strip()
                    has_keluar = attendance.get('jam_keluar') and attendance['jam_keluar'].strip()
                    if has_masuk or has_keluar:
                        total_present_on_date += 1
            
            total_item = QTableWidgetItem(str(total_present_on_date))
            total_item.setFlags(total_item.flags() & ~Qt.ItemIsEditable)
            total_item.setTextAlignment(Qt.AlignCenter)
            total_item.setBackground(QColor(240, 248, 255))
            total_item.setFont(QFont("", 0, QFont.Bold))
            
            # Highlight Sundays in summary row too
            if date.weekday() == 6:  # Sunday
                total_item.setBackground(QColor(255, 200, 200))
            
            self.table.setItem(summary_row, col, total_item)
        
        # Grand total (not really meaningful, so leave empty)
        grand_total_item = QTableWidgetItem("")
        grand_total_item.setFlags(grand_total_item.flags() & ~Qt.ItemIsEditable)
        grand_total_item.setBackground(QColor(240, 248, 255))
        self.table.setItem(summary_row, len(self.date_range) + 1, grand_total_item)
        
        # Adjust column widths
        self.table.setColumnWidth(0, 200)  # Name column wider
        for i in range(1, len(self.date_range) + 1):
            self.table.setColumnWidth(i, 80)  # Date columns
        self.table.setColumnWidth(len(self.date_range) + 1, 100)  # Total column
    
    def export_excel(self):
        """Export laporan ke Excel"""
        if not self.attendance_data:
            QMessageBox.warning(self, "Warning", "Tidak ada data untuk di-export. Generate laporan terlebih dahulu!")
            return
        
        # Get file path
        start_date = self.start_date.date().toPython()
        end_date = self.end_date.date().toPython()
        default_filename = f"Laporan_Kehadiran_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, 
            "Export Laporan Kehadiran", 
            default_filename,
            "Excel Files (*.xlsx)"
        )
        
        if not file_path:
            return
        
        try:
            # Show loading
            self.progress_bar.setVisible(True)
            self.progress_bar.setRange(0, 100)
            self.progress_bar.setValue(10)
            QApplication.processEvents()
            
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Laporan Kehadiran"
            
            # Title
            from openpyxl.utils import get_column_letter
            last_col = get_column_letter(len(self.date_range) + 2)  # +2 for name column and total column
            ws.merge_cells(f'A1:{last_col}1')
            ws['A1'] = "LAPORAN KEHADIRAN SEMUA KARYAWAN"
            ws['A1'].font = Font(size=16, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Period info
            period_text = f"Periode: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
            ws.merge_cells(f'A2:{last_col}2')
            ws['A2'] = period_text
            ws['A2'].font = Font(size=12, bold=True)
            ws['A2'].alignment = Alignment(horizontal='center')
            
            self.progress_bar.setValue(30)
            QApplication.processEvents()
            
            # Headers (starting from row 4)
            headers = ["Nama Karyawan"]
            for date in self.date_range:
                day_names = ["Sen", "Sel", "Rab", "Kam", "Jum", "Sab", "Min"]
                day_name = day_names[date.weekday()]
                date_str = f"{day_name}, {date.strftime('%d/%m')}"
                headers.append(date_str)
            headers.append("Total Hadir")
            
            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")
                
                # Highlight Sunday columns
                if col > 1 and col <= len(self.date_range) + 1:
                    date_idx = col - 2
                    if self.date_range[date_idx].weekday() == 6:  # Sunday
                        cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            
            self.progress_bar.setValue(50)
            QApplication.processEvents()
            
            # Data rows
            for row, employee in enumerate(self.employees, 5):
                # Employee name
                ws.cell(row=row, column=1).value = employee['name']
                
                # Attendance data
                total_present = 0
                for col, date in enumerate(self.date_range, 2):
                    date_str = date.strftime('%Y-%m-%d')
                    attendance = self.attendance_data[employee['id']].get(date_str)
                    
                    cell = ws.cell(row=row, column=col)
                    
                    # Check for leaves first
                    leaves = self.db_manager.get_leaves_by_employee_date(employee['id'], date_str)
                    has_leaves = len(leaves) > 0 if leaves else False
                    
                    if has_leaves:
                        # Employee has leave - green background with "Izin" text
                        if len(leaves) > 1:
                            cell.value = f"Izin ({len(leaves)})"
                        else:
                            cell.value = "Izin"
                        cell.fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")  # Light green
                        total_present += 1  # Count leave as present
                    elif attendance:
                        has_masuk = attendance.get('jam_masuk') and attendance['jam_masuk'].strip()
                        has_keluar = attendance.get('jam_keluar') and attendance['jam_keluar'].strip()
                        
                        if has_masuk and has_keluar:
                            # Complete data
                            cell.value = "✅"
                        elif has_masuk or has_keluar:
                            # Incomplete data - orange background
                            cell.value = "✅"
                            cell.fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
                        
                        if has_masuk or has_keluar:
                            total_present += 1
                    
                    # Highlight Sundays
                    if date.weekday() == 6:  # Sunday
                        if not cell.fill.start_color or cell.fill.start_color.rgb == "00000000":
                            cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                    
                    cell.alignment = Alignment(horizontal='center')
                
                # Total column
                total_cell = ws.cell(row=row, column=len(self.date_range) + 2)
                total_cell.value = total_present
                total_cell.alignment = Alignment(horizontal='center')
                total_cell.fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
            
            self.progress_bar.setValue(70)
            QApplication.processEvents()
            
            # Summary row
            summary_row = len(self.employees) + 5
            ws.cell(row=summary_row, column=1).value = "TOTAL HADIR"
            ws.cell(row=summary_row, column=1).font = Font(bold=True)
            ws.cell(row=summary_row, column=1).fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
            
            # Calculate totals per date
            for col, date in enumerate(self.date_range, 2):
                date_str = date.strftime('%Y-%m-%d')
                total_present_on_date = 0
                
                for employee in self.employees:
                    attendance = self.attendance_data[employee['id']].get(date_str)
                    if attendance:
                        has_masuk = attendance.get('jam_masuk') and attendance['jam_masuk'].strip()
                        has_keluar = attendance.get('jam_keluar') and attendance['jam_keluar'].strip()
                        if has_masuk or has_keluar:
                            total_present_on_date += 1
                
                cell = ws.cell(row=summary_row, column=col)
                cell.value = total_present_on_date
                cell.alignment = Alignment(horizontal='center')
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
                
                # Highlight Sundays in summary
                if date.weekday() == 6:  # Sunday
                    cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            
            self.progress_bar.setValue(90)
            QApplication.processEvents()
            
            # Legend
            legend_row = summary_row + 3
            ws.cell(row=legend_row, column=1).value = "KETERANGAN:"
            ws.cell(row=legend_row, column=1).font = Font(bold=True)
            
            ws.cell(row=legend_row + 1, column=1).value = "✅ = Hadir lengkap (jam masuk & keluar)"
            ws.cell(row=legend_row + 2, column=1).value = "✅ (orange) = Hadir tidak lengkap (salah satu jam kosong)"
            ws.cell(row=legend_row + 3, column=1).value = "Izin (hijau) = Karyawan izin"
            ws.cell(row=legend_row + 4, column=1).value = "(kosong) = Tidak hadir"
            ws.cell(row=legend_row + 5, column=1).value = "(merah) = Hari Minggu"
            
            # Auto-adjust column widths
            from openpyxl.utils import get_column_letter
            for col in range(1, len(headers) + 1):
                col_letter = get_column_letter(col)
                if col == 1:  # Name column
                    ws.column_dimensions[col_letter].width = 25
                elif col == len(headers):  # Total column
                    ws.column_dimensions[col_letter].width = 12
                else:  # Date columns
                    ws.column_dimensions[col_letter].width = 10
            
            # Save file
            wb.save(file_path)
            
            self.progress_bar.setValue(100)
            QApplication.processEvents()
            
            QMessageBox.information(self, "Success", f"Laporan berhasil di-export ke:\n{file_path}")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal export ke Excel: {str(e)}")
        finally:
            self.progress_bar.setVisible(False)


class LaporanPelanggaranSemuaDialog(QDialog):
    """Dialog untuk laporan pelanggaran semua karyawan"""
    def __init__(self, db_manager, parent=None):
        super().__init__(parent)
        self.db_manager = db_manager
        self.setWindowTitle("⚠️ Laporan Pelanggaran Semua Karyawan")
        self.setModal(True)
        self.resize(1400, 900)
        
        # Data storage
        self.violation_data = {}
        self.employees = []
        
        layout = QVBoxLayout()
        
        # Header
        header = QLabel("⚠️ LAPORAN PELANGGARAN & KETERLAMBATAN SEMUA KARYAWAN")
        header.setAlignment(Qt.AlignCenter)
        header.setStyleSheet("""
            QLabel {
                font-size: 20px;
                font-weight: bold;
                padding: 15px;
                background-color: #ffebee;
                border: 2px solid #e74c3c;
                border-radius: 8px;
                color: #c62828;
            }
        """)
        layout.addWidget(header)
        
        # Controls
        controls_layout = QHBoxLayout()
        
        controls_layout.addWidget(QLabel("Dari Tanggal:"))
        self.start_date = IndonesianDateEdit()
        self.start_date.setDate(QDate.currentDate().addDays(-30))
        controls_layout.addWidget(self.start_date)
        
        controls_layout.addWidget(QLabel("Sampai Tanggal:"))
        self.end_date = IndonesianDateEdit()
        self.end_date.setDate(QDate.currentDate())
        controls_layout.addWidget(self.end_date)
        
        generate_btn = QPushButton("🔄 Generate Laporan")
        generate_btn.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px 20px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
        """)
        generate_btn.clicked.connect(self.generate_report)
        controls_layout.addWidget(generate_btn)
        
        self.export_btn = QPushButton("📊 Export Excel")
        self.export_btn.setStyleSheet("""
            QPushButton {
                background-color: #28a745;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px 20px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #1e7e34;
            }
        """)
        self.export_btn.clicked.connect(self.export_excel)
        self.export_btn.setEnabled(False)
        controls_layout.addWidget(self.export_btn)
        
        controls_layout.addStretch()
        layout.addLayout(controls_layout)
        
        # Progress bar (hidden by default)
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        layout.addWidget(self.progress_bar)
        
        # Table with scroll area
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        
        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.horizontalHeader().setStretchLastSection(False)
        self.table.setStyleSheet("""
            QTableWidget {
                gridline-color: #dee2e6;
                background-color: white;
            }
            QTableWidget::item {
                padding: 8px;
                text-align: left;
            }
            QHeaderView::section {
                background-color: #f8f9fa;
                padding: 8px;
                border: 1px solid #dee2e6;
                font-weight: bold;
            }
        """)
        
        scroll_area.setWidget(self.table)
        layout.addWidget(scroll_area)
        
        # Summary info
        self.summary_label = QLabel("")
        self.summary_label.setStyleSheet("""
            QLabel {
                font-size: 12px;
                padding: 8px;
                background-color: #e9ecef;
                border-radius: 5px;
                color: #495057;
            }
        """)
        layout.addWidget(self.summary_label)
        
        # Close button
        close_btn = QPushButton("Tutup")
        close_btn.clicked.connect(self.close)
        close_layout = QHBoxLayout()
        close_layout.addStretch()
        close_layout.addWidget(close_btn)
        layout.addLayout(close_layout)
        
        self.setLayout(layout)
    
    def validate_date_range(self):
        """Validasi range tanggal"""
        start = self.start_date.date().toPython()
        end = self.end_date.date().toPython()
        
        if start > end:
            QMessageBox.warning(self, "Error", "Tanggal mulai tidak boleh lebih besar dari tanggal akhir!")
            return False
        
        # Check maksimal 3 bulan
        max_days = 90
        days_diff = (end - start).days + 1
        
        if days_diff > max_days:
            QMessageBox.warning(self, "Error", f"Range tanggal maksimal {max_days} hari!\nRange yang dipilih: {days_diff} hari")
            return False
        
        return True
    
    def generate_report(self):
        """Generate laporan pelanggaran semua karyawan"""
        if not self.validate_date_range():
            return
        
        # Show loading
        self.progress_bar.setVisible(True)
        self.progress_bar.setRange(0, 0)  # Indeterminate progress
        QApplication.processEvents()
        
        try:
            # Get date range
            start_date = self.start_date.date().toPython()
            end_date = self.end_date.date().toPython()
            
            # Get all employees (sorted alphabetically)
            self.employees = self.db_manager.get_all_employees()
            self.employees.sort(key=lambda x: x['name'])
            
            # Get violation data for all employees in date range
            self.violation_data = {}
            total_violations = 0
            total_violation_time = 0  # in minutes
            
            for employee in self.employees:
                # Get attendance records for this employee in the date range
                attendance_records = self.db_manager.get_attendance_by_employee_period(
                    employee['id'], 
                    start_date.strftime('%Y-%m-%d'), 
                    end_date.strftime('%Y-%m-%d')
                )
                
                # Get violations for each attendance record
                employee_violations = []
                for record in attendance_records:
                    if record.get('id'):
                        violations = self.db_manager.get_violations_by_attendance(record['id'])
                        for violation in violations:
                            # Calculate violation duration
                            duration_minutes = self.calculate_violation_duration(
                                violation['start_time'], violation['end_time']
                            )
                            
                            violation_info = {
                                'date': record['date'],
                                'description': violation['description'],
                                'start_time': violation['start_time'],
                                'end_time': violation['end_time'],
                                'duration_minutes': duration_minutes,
                                'duration_text': self.format_duration(duration_minutes)
                            }
                            employee_violations.append(violation_info)
                            total_violations += 1
                            total_violation_time += duration_minutes
                
                self.violation_data[employee['id']] = {
                    'name': employee['name'],
                    'violations': employee_violations,
                    'total_violations': len(employee_violations),
                    'total_time_minutes': sum(v['duration_minutes'] for v in employee_violations)
                }
            
            # Populate table
            self.populate_violation_table()
            
            # Update summary
            total_time_text = self.format_duration(total_violation_time)
            employees_with_violations = sum(1 for emp_data in self.violation_data.values() if emp_data['violations'])
            
            self.summary_label.setText(
                f"RINGKASAN: {employees_with_violations} karyawan memiliki pelanggaran | "
                f"Total {total_violations} pelanggaran | "
                f"Total waktu pelanggaran: {total_time_text} | "
                f"Periode: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
            )
            
            # Enable export button
            self.export_btn.setEnabled(True)
            
            if total_violations == 0:
                QMessageBox.information(self, "Info", "Tidak ada pelanggaran ditemukan dalam periode yang dipilih.")
            else:
                QMessageBox.information(self, "Success", 
                    f"Laporan berhasil dibuat!\n"
                    f"Periode: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}\n"
                    f"Total: {employees_with_violations} karyawan dengan pelanggaran\n"
                    f"Total pelanggaran: {total_violations}")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal membuat laporan: {str(e)}")
        finally:
            # Hide loading
            self.progress_bar.setVisible(False)
    
    def calculate_violation_duration(self, start_time, end_time):
        """Calculate duration in minutes between start_time and end_time"""
        try:
            from datetime import datetime
            
            # Parse times (format: HH:MM:SS)
            start = datetime.strptime(start_time, "%H:%M:%S")
            end = datetime.strptime(end_time, "%H:%M:%S")
            
            # Handle case where end time is next day (rare but possible)
            if end < start:
                end = end.replace(day=start.day + 1)
            
            # Calculate difference in minutes
            diff = end - start
            return int(diff.total_seconds() / 60)
        
        except Exception as e:
            print(f"Error calculating duration: {e}")
            return 0
    
    def format_duration(self, minutes):
        """Format duration in minutes to readable text"""
        if minutes == 0:
            return "0 menit"
        
        hours = minutes // 60
        mins = minutes % 60
        
        if hours > 0 and mins > 0:
            return f"{hours} jam {mins} menit"
        elif hours > 0:
            return f"{hours} jam"
        else:
            return f"{mins} menit"
    
    def populate_violation_table(self):
        """Populate tabel dengan format yang mudah dibaca berdasarkan karyawan"""
        if not self.violation_data:
            return
        
        # Setup table columns - format seperti pada gambar
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels([
            "NAMA", "TANGGAL", "HARI", "RENTANG WAKTU", "DURASI", "NOTE"
        ])
        
        # Calculate total rows needed
        total_rows = 0
        for emp_data in self.violation_data.values():
            if emp_data['violations']:
                total_rows += len(emp_data['violations'])  # Only violation rows
        
        # If no violations, show empty state
        if total_rows == 0:
            self.table.setRowCount(1)
            no_data_item = QTableWidgetItem("Tidak ada pelanggaran dalam periode yang dipilih")
            no_data_item.setTextAlignment(Qt.AlignCenter)
            no_data_item.setFont(QFont("", 0, QFont.Bold))
            no_data_item.setBackground(QColor(255, 248, 220))  # Light yellow
            self.table.setItem(0, 0, no_data_item)
            
            # Merge cells for the message
            self.table.setSpan(0, 0, 1, 6)
            return
        
        self.table.setRowCount(total_rows)
        current_row = 0
        
        # Populate data - group by employee
        for employee in self.employees:
            emp_data = self.violation_data[employee['id']]
            
            if not emp_data['violations']:
                continue  # Skip employees without violations
            
            # Sort violations by date
            violations_sorted = sorted(emp_data['violations'], key=lambda x: x['date'])
            
            for i, violation in enumerate(violations_sorted):
                # Employee name (only on first row for each employee)
                if i == 0:
                    name_item = QTableWidgetItem(emp_data['name'].upper())
                    name_item.setFont(QFont("", 0, QFont.Bold))
                    name_item.setBackground(QColor(173, 216, 230))  # Light blue like in image
                    name_item.setTextAlignment(Qt.AlignCenter)
                    name_item.setFlags(name_item.flags() & ~Qt.ItemIsEditable)
                    self.table.setItem(current_row, 0, name_item)
                    
                    # Merge cells for employee name if multiple violations
                    if len(violations_sorted) > 1:
                        self.table.setSpan(current_row, 0, len(violations_sorted), 1)
                else:
                    # Empty for subsequent rows (handled by span)
                    pass
                
                # Date
                date_item = QTableWidgetItem(violation['date'])
                date_item.setTextAlignment(Qt.AlignCenter)
                date_item.setFlags(date_item.flags() & ~Qt.ItemIsEditable)
                self.table.setItem(current_row, 1, date_item)
                
                # Day of week
                from datetime import datetime
                try:
                    date_obj = datetime.strptime(violation['date'], '%Y-%m-%d')
                    day_names = ['SEN', 'SEL', 'RAB', 'KAM', 'JUM', 'SAB', 'MIN']
                    day_name = day_names[date_obj.weekday()]
                except:
                    day_name = ""
                
                day_item = QTableWidgetItem(day_name)
                day_item.setTextAlignment(Qt.AlignCenter)
                day_item.setFlags(day_item.flags() & ~Qt.ItemIsEditable)
                self.table.setItem(current_row, 2, day_item)
                
                # Time range (separate column)
                time_range = f"{violation['start_time']} - {violation['end_time']}"
                time_range_item = QTableWidgetItem(time_range)
                time_range_item.setTextAlignment(Qt.AlignCenter)
                time_range_item.setFlags(time_range_item.flags() & ~Qt.ItemIsEditable)
                self.table.setItem(current_row, 3, time_range_item)
                
                # Duration (separate column)
                duration_text = f"{violation['duration_text']}"
                duration_item = QTableWidgetItem(duration_text)
                duration_item.setTextAlignment(Qt.AlignCenter)
                duration_item.setFlags(duration_item.flags() & ~Qt.ItemIsEditable)
                self.table.setItem(current_row, 4, duration_item)
                
                # Note (description)
                note_item = QTableWidgetItem(violation['description'].upper())
                note_item.setFlags(note_item.flags() & ~Qt.ItemIsEditable)
                self.table.setItem(current_row, 5, note_item)
                
                current_row += 1
        
        # Adjust column widths to match the format in image
        self.table.setColumnWidth(0, 120)  # NAMA
        self.table.setColumnWidth(1, 100)  # TANGGAL
        self.table.setColumnWidth(2, 60)   # HARI
        self.table.setColumnWidth(3, 150)  # RENTANG WAKTU
        self.table.setColumnWidth(4, 100)  # DURASI
        self.table.setColumnWidth(5, 300)  # NOTE
        
        # Set row height for better readability
        for row in range(self.table.rowCount()):
            self.table.setRowHeight(row, 45)
    
    def export_excel(self):
        """Export laporan pelanggaran ke Excel"""
        if not self.violation_data:
            QMessageBox.warning(self, "Warning", "Tidak ada data untuk di-export. Generate laporan terlebih dahulu!")
            return
        
        # Get file path
        start_date = self.start_date.date().toPython()
        end_date = self.end_date.date().toPython()
        default_filename = f"Laporan_Pelanggaran_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, 
            "Export Laporan Pelanggaran", 
            default_filename,
            "Excel Files (*.xlsx)"
        )
        
        if not file_path:
            return
        
        try:
            # Show loading
            self.progress_bar.setVisible(True)
            self.progress_bar.setRange(0, 100)
            self.progress_bar.setValue(10)
            QApplication.processEvents()
            
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Laporan Pelanggaran"
            
            # Title
            ws.merge_cells('A1:D1')
            ws['A1'] = "LAPORAN PELANGGARAN & KETERLAMBATAN SEMUA KARYAWAN"
            ws['A1'].font = Font(size=16, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Period info
            period_text = f"Periode: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
            ws.merge_cells('A2:D2')
            ws['A2'] = period_text
            ws['A2'].font = Font(size=12, bold=True)
            ws['A2'].alignment = Alignment(horizontal='center')
            
            self.progress_bar.setValue(30)
            QApplication.processEvents()
            
            # Headers (starting from row 4) - format seperti gambar
            headers = ["NAMA", "TANGGAL", "HARI", "RENTANG WAKTU", "DURASI", "NOTE"]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")
            
            self.progress_bar.setValue(50)
            QApplication.processEvents()
            
            # Data rows - format seperti tabel
            current_row = 5
            
            for employee in self.employees:
                emp_data = self.violation_data[employee['id']]
                
                if not emp_data['violations']:
                    continue  # Skip employees without violations
                
                # Sort violations by date
                violations_sorted = sorted(emp_data['violations'], key=lambda x: x['date'])
                
                for i, violation in enumerate(violations_sorted):
                    # Employee name (only on first row for each employee)
                    if i == 0:
                        name_cell = ws.cell(row=current_row, column=1)
                        name_cell.value = emp_data['name'].upper()
                        name_cell.font = Font(bold=True)
                        name_cell.fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Light blue
                        name_cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # Merge cells if multiple violations
                        if len(violations_sorted) > 1:
                            ws.merge_cells(start_row=current_row, start_column=1, 
                                         end_row=current_row + len(violations_sorted) - 1, end_column=1)
                    
                    # Date
                    date_cell = ws.cell(row=current_row, column=2)
                    date_cell.value = violation['date']
                    date_cell.alignment = Alignment(horizontal='center')
                    
                    # Day of week
                    from datetime import datetime
                    try:
                        date_obj = datetime.strptime(violation['date'], '%Y-%m-%d')
                        day_names = ['SEN', 'SEL', 'RAB', 'KAM', 'JUM', 'SAB', 'MIN']
                        day_name = day_names[date_obj.weekday()]
                    except:
                        day_name = ""
                    
                    day_cell = ws.cell(row=current_row, column=3)
                    day_cell.value = day_name
                    day_cell.alignment = Alignment(horizontal='center')
                    
                    # Time range (separate column)
                    time_range = f"{violation['start_time']} - {violation['end_time']}"
                    time_range_cell = ws.cell(row=current_row, column=4)
                    time_range_cell.value = time_range
                    time_range_cell.alignment = Alignment(horizontal='center')
                    
                    # Duration (separate column)
                    duration_text = f"{violation['duration_text']}"
                    duration_cell = ws.cell(row=current_row, column=5)
                    duration_cell.value = duration_text
                    duration_cell.alignment = Alignment(horizontal='center')
                    
                    # Note (description)
                    note_cell = ws.cell(row=current_row, column=6)
                    note_cell.value = violation['description'].upper()
                    
                    current_row += 1
            
            self.progress_bar.setValue(90)
            QApplication.processEvents()
            
            # Summary section
            summary_row = current_row + 2
            ws.cell(row=summary_row, column=1).value = "RINGKASAN:"
            ws.cell(row=summary_row, column=1).font = Font(bold=True)
            
            # Calculate summary data
            total_violations = sum(len(emp_data['violations']) for emp_data in self.violation_data.values())
            total_time_minutes = sum(emp_data['total_time_minutes'] for emp_data in self.violation_data.values())
            employees_with_violations = sum(1 for emp_data in self.violation_data.values() if emp_data['violations'])
            
            ws.cell(row=summary_row + 1, column=1).value = f"• Total karyawan dengan pelanggaran: {employees_with_violations}"
            ws.cell(row=summary_row + 2, column=1).value = f"• Total pelanggaran: {total_violations}"
            ws.cell(row=summary_row + 3, column=1).value = f"• Total waktu pelanggaran: {self.format_duration(total_time_minutes)}"
            
            # Auto-adjust column widths - sesuai format baru
            ws.column_dimensions['A'].width = 15  # NAMA
            ws.column_dimensions['B'].width = 12  # TANGGAL
            ws.column_dimensions['C'].width = 8   # HARI
            ws.column_dimensions['D'].width = 25  # WAKTU
            ws.column_dimensions['E'].width = 35  # NOTE
            
            # Save file
            wb.save(file_path)
            
            self.progress_bar.setValue(100)
            QApplication.processEvents()
            
            QMessageBox.information(self, "Success", f"Laporan berhasil di-export ke:\n{file_path}")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal export ke Excel: {str(e)}")
        finally:
            self.progress_bar.setVisible(False)


class LaporanOvertimeSemuaDialog(QDialog):
    """Dialog untuk laporan overtime semua karyawan"""
    def __init__(self, db_manager, parent=None):
        super().__init__(parent)
        self.db_manager = db_manager
        self.setWindowTitle("📈 Laporan Overtime Semua Karyawan")
        self.setModal(True)
        self.resize(1200, 800)
        
        layout = QVBoxLayout()
        
        # Header
        header = QLabel("📈 LAPORAN OVERTIME & LOYALITAS")
        header.setAlignment(Qt.AlignCenter)
        header.setStyleSheet("font-size: 18px; font-weight: bold; padding: 10px; color: #f39c12;")
        layout.addWidget(header)
        
        # Placeholder content
        content = QLabel("Fitur laporan overtime semua karyawan akan segera ditambahkan.\n\nFitur ini akan menampilkan:\n• Total overtime per karyawan\n• Total loyalitas per karyawan\n• Ranking karyawan berdasarkan overtime\n• Grafik trend overtime bulanan")
        content.setAlignment(Qt.AlignCenter)
        content.setStyleSheet("font-size: 14px; padding: 50px;")
        layout.addWidget(content)
        
        # Close button
        close_btn = QPushButton("Tutup")
        close_btn.clicked.connect(self.close)
        close_layout = QHBoxLayout()
        close_layout.addStretch()
        close_layout.addWidget(close_btn)
        layout.addLayout(close_layout)
        
        self.setLayout(layout)


class LaporanBulananDialog(QDialog):
    """Dialog untuk laporan bulanan"""
    def __init__(self, db_manager, parent=None):
        super().__init__(parent)
        self.db_manager = db_manager
        self.setWindowTitle("📅 Laporan Bulanan Rekap Absensi")
        self.setModal(True)
        self.resize(1200, 800)
        
        layout = QVBoxLayout()
        
        # Header
        header = QLabel("📅 REKAP ABSENSI BULANAN")
        header.setAlignment(Qt.AlignCenter)
        header.setStyleSheet("font-size: 18px; font-weight: bold; padding: 10px; color: #9b59b6;")
        layout.addWidget(header)
        
        # Placeholder content
        content = QLabel("Fitur laporan bulanan akan segera ditambahkan.\n\nFitur ini akan menampilkan:\n• Rekap kehadiran per bulan\n• Statistik keterlambatan\n• Total jam kerja dan overtime\n• Persentase kehadiran\n• Grafik trend bulanan")
        content.setAlignment(Qt.AlignCenter)
        content.setStyleSheet("font-size: 14px; padding: 50px;")
        layout.addWidget(content)
        
        # Close button
        close_btn = QPushButton("Tutup")
        close_btn.clicked.connect(self.close)
        close_layout = QHBoxLayout()
        close_layout.addStretch()
        close_layout.addWidget(close_btn)
        layout.addLayout(close_layout)
        
        self.setLayout(layout)


class LaporanKinerjaDialog(QDialog):
    """Dialog untuk laporan kinerja kehadiran"""
    def __init__(self, db_manager, parent=None):
        super().__init__(parent)
        self.db_manager = db_manager
        self.setWindowTitle("🏆 Laporan Kinerja Kehadiran")
        self.setModal(True)
        self.resize(1200, 800)
        
        layout = QVBoxLayout()
        
        # Header
        header = QLabel("🏆 ANALISIS KINERJA KEHADIRAN")
        header.setAlignment(Qt.AlignCenter)
        header.setStyleSheet("font-size: 18px; font-weight: bold; padding: 10px; color: #1abc9c;")
        layout.addWidget(header)
        
        # Placeholder content
        content = QLabel("Fitur laporan kinerja akan segera ditambahkan.\n\nFitur ini akan menampilkan:\n• Ranking karyawan terbaik\n• Skor kehadiran per karyawan\n• Analisis pola keterlambatan\n• Rekomendasi perbaikan\n• Dashboard kinerja visual")
        content.setAlignment(Qt.AlignCenter)
        content.setStyleSheet("font-size: 14px; padding: 50px;")
        layout.addWidget(content)
        
        # Close button
        close_btn = QPushButton("Tutup")
        close_btn.clicked.connect(self.close)
        close_layout = QHBoxLayout()
        close_layout.addStretch()
        close_layout.addWidget(close_btn)
        layout.addLayout(close_layout)
        
        self.setLayout(layout)


class ManagementTab(QWidget):
    """Tab Management untuk berbagai pengaturan sistem"""
    def __init__(self, db_manager, main_window=None):
        super().__init__()
        self.db_manager = db_manager
        self.main_window = main_window
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        
        # Header
        header_label = QLabel("⚙️ MANAGEMENT SISTEM")
        header_label.setAlignment(Qt.AlignCenter)
        header_label.setStyleSheet("""
            QLabel {
                font-size: 24px;
                font-weight: bold;
                color: #2c3e50;
                padding: 20px;
                background-color: #ecf0f1;
                border-radius: 10px;
                margin-bottom: 20px;
            }
        """)
        layout.addWidget(header_label)
        
        # Scroll area untuk menu buttons
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        
        # Widget container untuk buttons
        container = QWidget()
        grid_layout = QGridLayout(container)
        grid_layout.setSpacing(15)
        
        # Menu buttons
        self.create_management_buttons(grid_layout)
        
        scroll.setWidget(container)
        layout.addWidget(scroll)
        
        self.setLayout(layout)
    
    def create_management_buttons(self, layout):
        """Membuat menu buttons untuk management"""
        
        # Data menu buttons
        menu_items = [
            {
                'title': '🔄 Management Shift',
                'description': 'Kelola pengaturan shift kerja\ndan jadwal karyawan',
                'color': '#3498db',
                'action': self.open_shift_management
            },
            {
                'title': '👥 Management Karyawan',
                'description': 'Kelola data karyawan,\npelanggaran dan izin',
                'color': '#27ae60',
                'action': self.open_employee_management
            },
            {
                'title': '📧 Tambah Izin',
                'description': 'Tambah izin karyawan\ntanpa perlu input harian',
                'color': '#17a2b8',
                'action': self.open_add_izin_standalone
            },
            {
                'title': '🗄️ Backup Database',
                'description': 'Backup dan restore\ndata aplikasi',
                'color': '#e74c3c',
                'action': self.open_backup_management
            },
            {
                'title': '⚙️ Pengaturan Sistem',
                'description': 'Konfigurasi umum\naplikasi absensi',
                'color': '#f39c12',
                'action': self.open_system_settings
            }
        ]
        
        # Arrange buttons in grid (2 columns)
        row = 0
        col = 0
        for item in menu_items:
            btn = self.create_styled_button(
                item['title'], 
                item['description'], 
                item['color'],
                item['action']
            )
            layout.addWidget(btn, row, col)
            
            col += 1
            if col >= 2:  # 2 columns
                col = 0
                row += 1
    
    def create_styled_button(self, title, description, color, action):
        """Membuat button dengan style yang menarik"""
        btn = QPushButton()
        btn.setFixedSize(300, 120)
        btn.setCursor(Qt.PointingHandCursor)
        
        # Set text
        btn.setText(f"{title}\n\n{description}")
        
        # Set style
        btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {color};
                color: white;
                border: none;
                border-radius: 15px;
                font-size: 12px;
                font-weight: bold;
                text-align: center;
                padding: 10px;
            }}
            QPushButton:hover {{
                background-color: {self.darken_color(color)};
            }}
            QPushButton:pressed {{
                background-color: {self.darken_color(color, 0.8)};
            }}
        """)
        
        # Connect action
        btn.clicked.connect(action)
        
        return btn
    
    def darken_color(self, hex_color, factor=0.8):
        """Menggelapkan warna untuk efek hover"""
        hex_color = hex_color.lstrip('#')
        rgb = tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
        darkened = tuple(int(c * factor) for c in rgb)
        return f"#{darkened[0]:02x}{darkened[1]:02x}{darkened[2]:02x}"
    
    # Action methods untuk setiap menu
    def open_shift_management(self):
        """Buka management shift (existing ShiftManagementTab)"""
        dialog = ShiftManagementDialog(self.db_manager, self)
        dialog.exec()
    
    def open_employee_management(self):
        """Buka management karyawan"""
        dialog = EmployeeManagementDialog(self.db_manager, self)
        dialog.exec()
    
    def open_add_izin_standalone(self):
        """Buka form tambah izin standalone"""
        dialog = StandaloneAddLeaveDialog(self.db_manager, self)
        dialog.exec()
    
    def open_backup_management(self):
        """Buka backup management"""
        QMessageBox.information(self, "Info", "Fitur Backup Database akan segera ditambahkan!")
    
    def open_system_settings(self):
        """Buka pengaturan sistem"""
        QMessageBox.information(self, "Info", "Fitur Pengaturan Sistem akan segera ditambahkan!")


class ShiftManagementDialog(QDialog):
    """Dialog untuk management shift menggunakan ShiftManagementTab yang sudah ada"""
    def __init__(self, db_manager, parent=None):
        super().__init__(parent)
        self.db_manager = db_manager
        self.setWindowTitle("🔄 Management Shift")
        self.setModal(True)
        self.resize(1000, 700)
        
        layout = QVBoxLayout()
        
        # Embed existing ShiftManagementTab
        self.shift_tab = ShiftManagementTab(db_manager)
        layout.addWidget(self.shift_tab)
        
        # Close button
        close_btn = QPushButton("Tutup")
        close_btn.clicked.connect(self.close)
        close_layout = QHBoxLayout()
        close_layout.addStretch()
        close_layout.addWidget(close_btn)
        layout.addLayout(close_layout)
        
        self.setLayout(layout)


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.db_manager = DatabaseManager()
        self.init_ui()
    
    def init_ui(self):
        self.setWindowTitle("🏢 Aplikasi Absensi - Sistem Terpadu")
        self.setGeometry(100, 100, 1200, 800)
        
        # Set light theme styling
        self.setStyleSheet("""
            QMainWindow {
                background-color: #ffffff;
                color: #212529;
            }
            QWidget {
                background-color: #ffffff;
                color: #212529;
            }
            QTabWidget::pane {
                border: 2px solid #e9ecef;
                background-color: #ffffff;
                border-radius: 8px;
                margin-top: -1px;
            }
            QTabWidget::tab-bar {
                alignment: center;
            }
            QTabBar::tab {
                background-color: #f8f9fa;
                color: #495057;
                border: 2px solid #e9ecef;
                padding: 15px 25px;
                margin-right: 3px;
                border-top-left-radius: 10px;
                border-top-right-radius: 10px;
                font-weight: bold;
                font-size: 14px;
                min-width: 120px;
            }
            QTabBar::tab:selected {
                background-color: #007bff;
                color: white;
                border-bottom: 2px solid #ffffff;
                margin-bottom: -2px;
            }
            QTabBar::tab:hover:!selected {
                background-color: #e2e6ea;
                color: #495057;
            }
            QTableWidget {
                background-color: #ffffff;
                alternate-background-color: #f8f9fa;
                gridline-color: #dee2e6;
                color: #212529;
                border: 1px solid #dee2e6;
                border-radius: 5px;
            }
            QTableWidget::item {
                padding: 8px;
                border-bottom: 1px solid #dee2e6;
            }
            QTableWidget::item:selected {
                background-color: #cce7ff;
                color: #212529;
            }
            QPushButton {
                background-color: #f8f9fa;
                color: #495057;
                border: 1px solid #ced4da;
                border-radius: 5px;
                padding: 8px 16px;
                font-weight: 500;
            }
            QPushButton:hover {
                background-color: #e2e6ea;
                border-color: #adb5bd;
            }
            QPushButton:pressed {
                background-color: #d1ecf1;
                border-color: #bee5eb;
            }
            QGroupBox {
                font-weight: bold;
                border: 2px solid #dee2e6;
                border-radius: 8px;
                margin-top: 1ex;
                background-color: #ffffff;
                color: #495057;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 10px 0 10px;
                background-color: #ffffff;
                color: #495057;
            }
            QLabel {
                color: #495057;
                background-color: transparent;
            }
            QComboBox, QDateEdit, QTimeEdit, QLineEdit, QTextEdit {
                background-color: #ffffff;
                color: #495057;
                border: 1px solid #ced4da;
                border-radius: 4px;
                padding: 6px;
            }
            QComboBox:focus, QDateEdit:focus, QTimeEdit:focus, QLineEdit:focus, QTextEdit:focus {
                border-color: #80bdff;
                outline: 0;
            }
        """)
        
        # Create tab widget
        self.tab_widget = QTabWidget()
        
        # Add new tabs with improved structure
        self.laporan_tab = LaporanTab(self.db_manager, self)
        self.attendance_tab = AttendanceInputTab(self.db_manager, self)
        self.management_tab = ManagementTab(self.db_manager, self)
        
        # Add tabs with icons and better names
        self.tab_widget.addTab(self.laporan_tab, "📊 Laporan")
        self.tab_widget.addTab(self.attendance_tab, "📝 Input Harian")
        self.tab_widget.addTab(self.management_tab, "⚙️ Management")
        
        self.setCentralWidget(self.tab_widget)
    
    def refresh_report_tab(self):
        """Refresh report tab setelah data baru disimpan"""
        # Note: Laporan tab tidak perlu refresh karena menggunakan dialog
        # yang akan refresh otomatis saat dibuka
        pass

def main():
    app = QApplication(sys.argv)
    
    # Set application style to light theme
    app.setStyle('Windows')  # Use Windows style for light theme
    
    # Set light theme palette
    from PySide6.QtGui import QPalette
    palette = QPalette()
    
    # Light theme colors
    palette.setColor(QPalette.Window, QColor(255, 255, 255))          # White background
    palette.setColor(QPalette.WindowText, QColor(0, 0, 0))           # Black text
    palette.setColor(QPalette.Base, QColor(255, 255, 255))           # White input background
    palette.setColor(QPalette.AlternateBase, QColor(245, 245, 245))  # Light gray alternate
    palette.setColor(QPalette.ToolTipBase, QColor(255, 255, 220))    # Light yellow tooltip
    palette.setColor(QPalette.ToolTipText, QColor(0, 0, 0))          # Black tooltip text
    palette.setColor(QPalette.Text, QColor(0, 0, 0))                 # Black text
    palette.setColor(QPalette.Button, QColor(240, 240, 240))         # Light gray button
    palette.setColor(QPalette.ButtonText, QColor(0, 0, 0))           # Black button text
    palette.setColor(QPalette.BrightText, QColor(255, 0, 0))         # Red bright text
    palette.setColor(QPalette.Link, QColor(42, 130, 218))            # Blue links
    palette.setColor(QPalette.Highlight, QColor(42, 130, 218))       # Blue selection
    palette.setColor(QPalette.HighlightedText, QColor(255, 255, 255)) # White selected text
    
    app.setPalette(palette)
    
    window = MainWindow()
    window.show()
    
    sys.exit(app.exec())

class StandaloneAddLeaveDialog(QDialog):
    """Dialog untuk menambah izin tanpa perlu entry di Input Harian"""
    def __init__(self, db_manager, parent=None):
        super().__init__(parent)
        self.db_manager = db_manager
        self.setWindowTitle("📧 Tambah Izin Karyawan")
        self.setModal(True)
        self.resize(450, 300)
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        
        # Header
        header = QLabel("📧 TAMBAH IZIN KARYAWAN")
        header.setAlignment(Qt.AlignCenter)
        header.setStyleSheet("""
            QLabel {
                font-size: 18px;
                font-weight: bold;
                padding: 15px;
                background-color: #17a2b8;
                color: white;
                border-radius: 8px;
                margin-bottom: 20px;
            }
        """)
        layout.addWidget(header)
        
        # Form
        form_layout = QFormLayout()
        
        # Employee selection
        self.employee_combo = QComboBox()
        employees = self.db_manager.get_all_employees()
        for emp in employees:
            self.employee_combo.addItem(emp['name'], emp['id'])
        form_layout.addRow("Nama Karyawan:", self.employee_combo)
        
        # Date selection
        self.date_edit = IndonesianDateEdit()
        self.date_edit.setDate(QDate.currentDate())
        form_layout.addRow("Tanggal Izin:", self.date_edit)
        
        # Description
        self.description_edit = QTextEdit()
        self.description_edit.setMaximumHeight(100)
        self.description_edit.setPlaceholderText("Masukkan keterangan izin (contoh: Sakit, Urusan keluarga, dll)")
        form_layout.addRow("Keterangan:", self.description_edit)
        
        layout.addLayout(form_layout)
        
        # Buttons
        button_layout = QHBoxLayout()
        
        save_btn = QPushButton("💾 Simpan Izin")
        save_btn.setStyleSheet("""
            QPushButton {
                background-color: #28a745;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px 20px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #218838;
            }
        """)
        save_btn.clicked.connect(self.save_leave)
        
        cancel_btn = QPushButton("❌ Batal")
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #6c757d;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px 20px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #5a6268;
            }
        """)
        cancel_btn.clicked.connect(self.reject)
        
        button_layout.addWidget(save_btn)
        button_layout.addWidget(cancel_btn)
        layout.addLayout(button_layout)
        
        self.setLayout(layout)
    
    def save_leave(self):
        """Simpan data izin"""
        if not self.description_edit.toPlainText().strip():
            QMessageBox.warning(self, "Error", "Keterangan izin tidak boleh kosong!")
            return
        
        employee_id = self.employee_combo.currentData()
        date = self.date_edit.date().toPython().strftime('%Y-%m-%d')
        description = self.description_edit.toPlainText().strip()
        
        try:
            self.db_manager.add_leave(employee_id, date, description)
            QMessageBox.information(self, "Sukses", "Izin berhasil ditambahkan!")
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal menyimpan izin: {str(e)}")


class EmployeeManagementDialog(QDialog):
    """Dialog untuk management karyawan dengan kelola pelanggaran dan izin"""
    def __init__(self, db_manager, parent=None):
        super().__init__(parent)
        self.db_manager = db_manager
        self.setWindowTitle("👥 Management Karyawan")
        self.setModal(True)
        self.resize(1000, 700)
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        
        # Header
        header = QLabel("👥 MANAGEMENT KARYAWAN")
        header.setAlignment(Qt.AlignCenter)
        header.setStyleSheet("""
            QLabel {
                font-size: 20px;
                font-weight: bold;
                padding: 15px;
                background-color: #27ae60;
                color: white;
                border-radius: 8px;
                margin-bottom: 20px;
            }
        """)
        layout.addWidget(header)
        
        # Employee list
        self.employee_table = QTableWidget()
        self.employee_table.setColumnCount(4)
        self.employee_table.setHorizontalHeaderLabels(["Nama Karyawan", "Kelola Pelanggaran", "Kelola Izin", "Total Pelanggaran/Izin"])
        self.employee_table.horizontalHeader().setStretchLastSection(True)
        self.employee_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.employee_table.setAlternatingRowColors(True)
        self.employee_table.verticalHeader().setDefaultSectionSize(50)
        
        # Style table
        self.employee_table.setStyleSheet("""
            QTableWidget {
                gridline-color: #dee2e6;
                background-color: white;
            }
            QTableWidget::item {
                padding: 8px;
                border-bottom: 1px solid #dee2e6;
            }
            QHeaderView::section {
                background-color: #f8f9fa;
                padding: 8px;
                border: 1px solid #dee2e6;
                font-weight: bold;
            }
        """)
        
        layout.addWidget(self.employee_table)
        
        # Buttons
        button_layout = QHBoxLayout()
        
        refresh_btn = QPushButton("🔄 Refresh")
        refresh_btn.clicked.connect(self.populate_employee_table)
        
        close_btn = QPushButton("❌ Tutup")
        close_btn.clicked.connect(self.close)
        
        button_layout.addStretch()
        button_layout.addWidget(refresh_btn)
        button_layout.addWidget(close_btn)
        layout.addLayout(button_layout)
        
        self.setLayout(layout)
        
        # Populate table
        self.populate_employee_table()
    
    def populate_employee_table(self):
        """Populate tabel karyawan"""
        employees = self.db_manager.get_all_employees()
        self.employee_table.setRowCount(len(employees))
        
        for row, employee in enumerate(employees):
            # Employee name
            name_item = QTableWidgetItem(employee['name'])
            name_item.setFlags(name_item.flags() & ~Qt.ItemIsEditable)
            self.employee_table.setItem(row, 0, name_item)
            
            # Kelola Pelanggaran button
            violation_btn = QPushButton("⚠️ Kelola Pelanggaran")
            violation_btn.setStyleSheet("""
                QPushButton {
                    background-color: #dc3545;
                    color: white;
                    border: none;
                    border-radius: 5px;
                    padding: 8px;
                    font-size: 11px;
                }
                QPushButton:hover {
                    background-color: #c82333;
                }
            """)
            violation_btn.clicked.connect(lambda checked, emp_id=employee['id'], emp_name=employee['name']: 
                                        self.manage_employee_violations(emp_id, emp_name))
            self.employee_table.setCellWidget(row, 1, violation_btn)
            
            # Kelola Izin button
            leave_btn = QPushButton("📧 Kelola Izin")
            leave_btn.setStyleSheet("""
                QPushButton {
                    background-color: #28a745;
                    color: white;
                    border: none;
                    border-radius: 5px;
                    padding: 8px;
                    font-size: 11px;
                }
                QPushButton:hover {
                    background-color: #218838;
                }
            """)
            leave_btn.clicked.connect(lambda checked, emp_id=employee['id'], emp_name=employee['name']: 
                                    self.manage_employee_leaves(emp_id, emp_name))
            self.employee_table.setCellWidget(row, 2, leave_btn)
            
            # Count violations and leaves
            # Note: We need to get all attendance records and count violations
            # For now, show placeholder
            count_item = QTableWidgetItem("Lihat detail →")
            count_item.setFlags(count_item.flags() & ~Qt.ItemIsEditable)
            count_item.setTextAlignment(Qt.AlignCenter)
            self.employee_table.setItem(row, 3, count_item)
        
        # Adjust column widths
        self.employee_table.setColumnWidth(0, 200)
        self.employee_table.setColumnWidth(1, 150)
        self.employee_table.setColumnWidth(2, 150)
    
    def manage_employee_violations(self, employee_id, employee_name):
        """Kelola pelanggaran karyawan"""
        dialog = EmployeeViolationManagementDialog(self.db_manager, employee_id, employee_name, self)
        dialog.exec()
    
    def manage_employee_leaves(self, employee_id, employee_name):
        """Kelola izin karyawan"""
        dialog = EmployeeLeaveManagementDialog(self.db_manager, employee_id, employee_name, self)
        dialog.exec()


class EmployeeViolationManagementDialog(QDialog):
    """Dialog untuk kelola pelanggaran per karyawan"""
    def __init__(self, db_manager, employee_id, employee_name, parent=None):
        super().__init__(parent)
        self.db_manager = db_manager
        self.employee_id = employee_id
        self.employee_name = employee_name
        self.setWindowTitle(f"⚠️ Pelanggaran - {employee_name}")
        self.setModal(True)
        self.resize(800, 600)
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        
        # Header
        header = QLabel(f"⚠️ PELANGGARAN - {self.employee_name.upper()}")
        header.setAlignment(Qt.AlignCenter)
        header.setStyleSheet("""
            QLabel {
                font-size: 16px;
                font-weight: bold;
                padding: 15px;
                background-color: #dc3545;
                color: white;
                border-radius: 8px;
                margin-bottom: 20px;
            }
        """)
        layout.addWidget(header)
        
        # Info
        info_label = QLabel("Menampilkan semua pelanggaran karyawan dari semua tanggal")
        info_label.setStyleSheet("color: #6c757d; font-style: italic; margin-bottom: 10px;")
        layout.addWidget(info_label)
        
        # Table
        self.violation_table = QTableWidget()
        self.violation_table.setColumnCount(5)
        self.violation_table.setHorizontalHeaderLabels(["Tanggal", "Jam Mulai", "Jam Selesai", "Keterangan", "Aksi"])
        self.violation_table.horizontalHeader().setStretchLastSection(True)
        self.violation_table.setAlternatingRowColors(True)
        self.violation_table.verticalHeader().setDefaultSectionSize(45)
        
        layout.addWidget(self.violation_table)
        
        # Buttons
        button_layout = QHBoxLayout()
        
        add_btn = QPushButton("➕ Tambah Pelanggaran")
        add_btn.setStyleSheet("""
            QPushButton {
                background-color: #dc3545;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px 20px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #c82333;
            }
        """)
        add_btn.clicked.connect(self.add_violation)
        
        close_btn = QPushButton("❌ Tutup")
        close_btn.clicked.connect(self.close)
        
        button_layout.addWidget(add_btn)
        button_layout.addStretch()
        button_layout.addWidget(close_btn)
        layout.addLayout(button_layout)
        
        self.setLayout(layout)
        
        # Populate table
        self.populate_violation_table()
    
    def populate_violation_table(self):
        """Populate tabel pelanggaran"""
        # Get all attendance records for this employee
        all_attendance = self.db_manager.get_attendance_by_employee_period(
            self.employee_id, '2020-01-01', '2030-12-31'  # Wide date range
        )
        
        violations = []
        for attendance in all_attendance:
            if 'id' in attendance and attendance['id']:
                emp_violations = self.db_manager.get_violations_by_attendance(attendance['id'])
                for violation in emp_violations:
                    violation['date'] = attendance['date']
                    violations.append(violation)
        
        self.violation_table.setRowCount(len(violations))
        
        for row, violation in enumerate(violations):
            # Date
            date_item = QTableWidgetItem(violation['date'])
            date_item.setFlags(date_item.flags() & ~Qt.ItemIsEditable)
            self.violation_table.setItem(row, 0, date_item)
            
            # Start time
            start_item = QTableWidgetItem(violation['start_time'])
            start_item.setFlags(start_item.flags() & ~Qt.ItemIsEditable)
            self.violation_table.setItem(row, 1, start_item)
            
            # End time
            end_item = QTableWidgetItem(violation['end_time'])
            end_item.setFlags(end_item.flags() & ~Qt.ItemIsEditable)
            self.violation_table.setItem(row, 2, end_item)
            
            # Description
            desc_item = QTableWidgetItem(violation['description'])
            desc_item.setFlags(desc_item.flags() & ~Qt.ItemIsEditable)
            self.violation_table.setItem(row, 3, desc_item)
            
            # Action buttons
            action_widget = QWidget()
            action_layout = QHBoxLayout(action_widget)
            action_layout.setContentsMargins(2, 2, 2, 2)
            action_layout.setSpacing(2)
            
            edit_btn = QPushButton("📝")
            edit_btn.setFixedSize(30, 25)
            edit_btn.setToolTip("Edit")
            edit_btn.setStyleSheet("""
                QPushButton {
                    background-color: #ffc107;
                    border: none;
                    border-radius: 3px;
                    font-size: 12px;
                    color: black;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #e0a800;
                }
            """)
            edit_btn.clicked.connect(lambda checked, v_id=violation['id']: self.edit_violation(v_id))
            
            delete_btn = QPushButton("❌")
            delete_btn.setFixedSize(30, 25)
            delete_btn.setToolTip("Hapus")
            delete_btn.setStyleSheet("""
                QPushButton {
                    background-color: #dc3545;
                    border: none;
                    border-radius: 3px;
                    font-size: 12px;
                    color: white;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #c82333;
                }
            """)
            delete_btn.clicked.connect(lambda checked, v_id=violation['id']: self.delete_violation(v_id))
            
            action_layout.addWidget(edit_btn)
            action_layout.addWidget(delete_btn)
            action_layout.addStretch()
            
            self.violation_table.setCellWidget(row, 4, action_widget)
    
    def add_violation(self):
        """Tambah pelanggaran baru"""
        dialog = AddViolationForEmployeeDialog(self.db_manager, self.employee_id, self.employee_name, self)
        if dialog.exec() == QDialog.Accepted:
            self.populate_violation_table()
    
    def edit_violation(self, violation_id):
        """Edit pelanggaran"""
        QMessageBox.information(self, "Info", f"Edit pelanggaran ID: {violation_id}")
    
    def delete_violation(self, violation_id):
        """Hapus pelanggaran"""
        reply = QMessageBox.question(self, "Konfirmasi", "Yakin ingin menghapus pelanggaran ini?")
        if reply == QMessageBox.Yes:
            try:
                self.db_manager.delete_violation(violation_id)
                QMessageBox.information(self, "Sukses", "Pelanggaran berhasil dihapus!")
                self.populate_violation_table()
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Gagal menghapus pelanggaran: {str(e)}")


class EmployeeLeaveManagementDialog(QDialog):
    """Dialog untuk kelola izin per karyawan"""
    def __init__(self, db_manager, employee_id, employee_name, parent=None):
        super().__init__(parent)
        self.db_manager = db_manager
        self.employee_id = employee_id
        self.employee_name = employee_name
        self.setWindowTitle(f"📧 Izin - {employee_name}")
        self.setModal(True)
        self.resize(700, 500)
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        
        # Header
        header = QLabel(f"📧 IZIN - {self.employee_name.upper()}")
        header.setAlignment(Qt.AlignCenter)
        header.setStyleSheet("""
            QLabel {
                font-size: 16px;
                font-weight: bold;
                padding: 15px;
                background-color: #28a745;
                color: white;
                border-radius: 8px;
                margin-bottom: 20px;
            }
        """)
        layout.addWidget(header)
        
        # Table
        self.leave_table = QTableWidget()
        self.leave_table.setColumnCount(4)
        self.leave_table.setHorizontalHeaderLabels(["Tanggal", "Keterangan", "Dibuat", "Aksi"])
        self.leave_table.horizontalHeader().setStretchLastSection(True)
        self.leave_table.setAlternatingRowColors(True)
        self.leave_table.verticalHeader().setDefaultSectionSize(45)
        
        layout.addWidget(self.leave_table)
        
        # Buttons
        button_layout = QHBoxLayout()
        
        add_btn = QPushButton("➕ Tambah Izin")
        add_btn.setStyleSheet("""
            QPushButton {
                background-color: #28a745;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px 20px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #218838;
            }
        """)
        add_btn.clicked.connect(self.add_leave)
        
        close_btn = QPushButton("❌ Tutup")
        close_btn.clicked.connect(self.close)
        
        button_layout.addWidget(add_btn)
        button_layout.addStretch()
        button_layout.addWidget(close_btn)
        layout.addLayout(button_layout)
        
        self.setLayout(layout)
        
        # Populate table
        self.populate_leave_table()
    
    def populate_leave_table(self):
        """Populate tabel izin"""
        # Get all leaves for this employee
        leaves = self.db_manager.get_leaves_by_employee_date(self.employee_id, None)  # Get all dates
        if not leaves:
            leaves = []
        
        self.leave_table.setRowCount(len(leaves))
        
        for row, leave in enumerate(leaves):
            # Date
            date_item = QTableWidgetItem(leave['date'])
            date_item.setFlags(date_item.flags() & ~Qt.ItemIsEditable)
            self.leave_table.setItem(row, 0, date_item)
            
            # Description
            desc_item = QTableWidgetItem(leave['description'])
            desc_item.setFlags(desc_item.flags() & ~Qt.ItemIsEditable)
            self.leave_table.setItem(row, 1, desc_item)
            
            # Created at
            created_item = QTableWidgetItem(leave.get('created_at', '-'))
            created_item.setFlags(created_item.flags() & ~Qt.ItemIsEditable)
            self.leave_table.setItem(row, 2, created_item)
            
            # Action buttons
            action_widget = QWidget()
            action_layout = QHBoxLayout(action_widget)
            action_layout.setContentsMargins(2, 2, 2, 2)
            action_layout.setSpacing(2)
            
            edit_btn = QPushButton("📝")
            edit_btn.setFixedSize(30, 25)
            edit_btn.setToolTip("Edit")
            edit_btn.setStyleSheet("""
                QPushButton {
                    background-color: #ffc107;
                    border: none;
                    border-radius: 3px;
                    font-size: 12px;
                    color: black;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #e0a800;
                }
            """)
            edit_btn.clicked.connect(lambda checked, l_id=leave['id']: self.edit_leave(l_id))
            
            delete_btn = QPushButton("❌")
            delete_btn.setFixedSize(30, 25)
            delete_btn.setToolTip("Hapus")
            delete_btn.setStyleSheet("""
                QPushButton {
                    background-color: #dc3545;
                    border: none;
                    border-radius: 3px;
                    font-size: 12px;
                    color: white;
                    font-weight: bold;
                }
                QPushButton:hover {
                    background-color: #c82333;
                }
            """)
            delete_btn.clicked.connect(lambda checked, l_id=leave['id']: self.delete_leave(l_id))
            
            action_layout.addWidget(edit_btn)
            action_layout.addWidget(delete_btn)
            action_layout.addStretch()
            
            self.leave_table.setCellWidget(row, 3, action_widget)
    
    def add_leave(self):
        """Tambah izin baru"""
        dialog = AddLeaveForEmployeeDialog(self.db_manager, self.employee_id, self.employee_name, self)
        if dialog.exec() == QDialog.Accepted:
            self.populate_leave_table()
    
    def edit_leave(self, leave_id):
        """Edit izin"""
        # Get leave data
        leaves = self.db_manager.get_leaves_by_employee_date(self.employee_id, None)
        leave_data = next((l for l in leaves if l['id'] == leave_id), None)
        
        if leave_data:
            dialog = EditLeaveDialog(self.db_manager, leave_data, self)
            if dialog.exec() == QDialog.Accepted:
                self.populate_leave_table()
    
    def delete_leave(self, leave_id):
        """Hapus izin"""
        reply = QMessageBox.question(self, "Konfirmasi", "Yakin ingin menghapus izin ini?")
        if reply == QMessageBox.Yes:
            try:
                self.db_manager.delete_leave(leave_id)
                QMessageBox.information(self, "Sukses", "Izin berhasil dihapus!")
                self.populate_leave_table()
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Gagal menghapus izin: {str(e)}")


class AddLeaveForEmployeeDialog(QDialog):
    """Dialog untuk tambah izin untuk karyawan tertentu"""
    def __init__(self, db_manager, employee_id, employee_name, parent=None):
        super().__init__(parent)
        self.db_manager = db_manager
        self.employee_id = employee_id
        self.employee_name = employee_name
        self.setWindowTitle(f"📧 Tambah Izin - {employee_name}")
        self.setModal(True)
        self.resize(400, 250)
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        
        # Header
        header = QLabel(f"📧 TAMBAH IZIN\n{self.employee_name}")
        header.setAlignment(Qt.AlignCenter)
        header.setStyleSheet("""
            QLabel {
                font-size: 14px;
                font-weight: bold;
                padding: 15px;
                background-color: #28a745;
                color: white;
                border-radius: 8px;
                margin-bottom: 20px;
            }
        """)
        layout.addWidget(header)
        
        # Form
        form_layout = QFormLayout()
        
        # Date selection
        self.date_edit = IndonesianDateEdit()
        self.date_edit.setDate(QDate.currentDate())
        form_layout.addRow("Tanggal Izin:", self.date_edit)
        
        # Description
        self.description_edit = QTextEdit()
        self.description_edit.setMaximumHeight(80)
        self.description_edit.setPlaceholderText("Masukkan keterangan izin")
        form_layout.addRow("Keterangan:", self.description_edit)
        
        layout.addLayout(form_layout)
        
        # Buttons
        button_layout = QHBoxLayout()
        
        save_btn = QPushButton("💾 Simpan")
        save_btn.clicked.connect(self.save_leave)
        
        cancel_btn = QPushButton("❌ Batal")
        cancel_btn.clicked.connect(self.reject)
        
        button_layout.addWidget(save_btn)
        button_layout.addWidget(cancel_btn)
        layout.addLayout(button_layout)
        
        self.setLayout(layout)
    
    def save_leave(self):
        """Simpan izin"""
        if not self.description_edit.toPlainText().strip():
            QMessageBox.warning(self, "Error", "Keterangan izin tidak boleh kosong!")
            return
        
        date = self.date_edit.date().toPython().strftime('%Y-%m-%d')
        description = self.description_edit.toPlainText().strip()
        
        try:
            self.db_manager.add_leave(self.employee_id, date, description)
            QMessageBox.information(self, "Sukses", "Izin berhasil ditambahkan!")
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal menyimpan izin: {str(e)}")


class AddViolationForEmployeeDialog(QDialog):
    """Dialog untuk tambah pelanggaran untuk karyawan tertentu"""
    def __init__(self, db_manager, employee_id, employee_name, parent=None):
        super().__init__(parent)
        self.db_manager = db_manager
        self.employee_id = employee_id
        self.employee_name = employee_name
        self.setWindowTitle(f"⚠️ Tambah Pelanggaran - {employee_name}")
        self.setModal(True)
        self.resize(500, 400)
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        
        # Header
        header = QLabel(f"⚠️ TAMBAH PELANGGARAN\n{self.employee_name}")
        header.setAlignment(Qt.AlignCenter)
        header.setStyleSheet("""
            QLabel {
                font-size: 14px;
                font-weight: bold;
                padding: 15px;
                background-color: #dc3545;
                color: white;
                border-radius: 8px;
                margin-bottom: 20px;
            }
        """)
        layout.addWidget(header)
        
        # Info
        info_label = QLabel("Catatan: Pelanggaran akan ditambahkan ke data kehadiran. Jika belum ada data kehadiran untuk tanggal tersebut, akan dibuat otomatis.")
        info_label.setStyleSheet("""
            QLabel {
                color: #6c757d;
                font-style: italic;
                padding: 10px;
                background-color: #f8f9fa;
                border-radius: 5px;
                margin-bottom: 15px;
            }
        """)
        info_label.setWordWrap(True)
        layout.addWidget(info_label)
        
        # Form
        form_layout = QFormLayout()
        
        # Date selection
        self.date_edit = IndonesianDateEdit()
        self.date_edit.setDate(QDate.currentDate())
        form_layout.addRow("Tanggal:", self.date_edit)
        
        # Start time
        self.start_time_edit = QTimeEdit()
        self.start_time_edit.setTime(QTime(9, 0))
        self.start_time_edit.setDisplayFormat("HH:mm")
        form_layout.addRow("Jam Mulai:", self.start_time_edit)
        
        # End time
        self.end_time_edit = QTimeEdit()
        self.end_time_edit.setTime(QTime(10, 0))
        self.end_time_edit.setDisplayFormat("HH:mm")
        form_layout.addRow("Jam Selesai:", self.end_time_edit)
        
        # Description
        self.description_edit = QTextEdit()
        self.description_edit.setMaximumHeight(100)
        self.description_edit.setPlaceholderText("Masukkan keterangan pelanggaran (contoh: Tidur, Makan di luar jam istirahat, dll)")
        form_layout.addRow("Keterangan:", self.description_edit)
        
        layout.addLayout(form_layout)
        
        # Buttons
        button_layout = QHBoxLayout()
        
        save_btn = QPushButton("💾 Simpan Pelanggaran")
        save_btn.setStyleSheet("""
            QPushButton {
                background-color: #dc3545;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px 20px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #c82333;
            }
        """)
        save_btn.clicked.connect(self.save_violation)
        
        cancel_btn = QPushButton("❌ Batal")
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #6c757d;
                color: white;
                border: none;
                border-radius: 5px;
                padding: 10px 20px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #5a6268;
            }
        """)
        cancel_btn.clicked.connect(self.reject)
        
        button_layout.addWidget(save_btn)
        button_layout.addWidget(cancel_btn)
        layout.addLayout(button_layout)
        
        self.setLayout(layout)
    
    def save_violation(self):
        """Simpan data pelanggaran"""
        if not self.description_edit.toPlainText().strip():
            QMessageBox.warning(self, "Error", "Keterangan pelanggaran tidak boleh kosong!")
            return
        
        # Validate time
        start_time = self.start_time_edit.time()
        end_time = self.end_time_edit.time()
        
        if start_time >= end_time:
            QMessageBox.warning(self, "Error", "Jam mulai harus lebih kecil dari jam selesai!")
            return
        
        date = self.date_edit.date().toPython().strftime('%Y-%m-%d')
        start_time_str = start_time.toString("HH:mm:ss")
        end_time_str = end_time.toString("HH:mm:ss")
        description = self.description_edit.toPlainText().strip()
        
        try:
            # Check if attendance record exists for this date
            attendance_data = self.db_manager.get_attendance_by_date(date)
            employee_attendance = None
            
            for record in attendance_data:
                if record['employee_id'] == self.employee_id:
                    employee_attendance = record
                    break
            
            # If no attendance record exists, create one
            if not employee_attendance:
                # Create minimal attendance record
                attendance_record = {
                    'employee_id': self.employee_id,
                    'date': date,
                    'jam_masuk': None,
                    'jam_keluar': None,
                    'jam_masuk_lembur': None,
                    'jam_keluar_lembur': None,
                    'shift_id': 1,  # Default shift
                    'keterangan': f"Data dibuat untuk pelanggaran: {description}"
                }
                
                # Save attendance record
                attendance_id = self.db_manager.save_attendance_data(date, [attendance_record])
                if attendance_id:
                    attendance_id = attendance_id[0] if isinstance(attendance_id, list) else attendance_id
                else:
                    # Get the created attendance ID
                    updated_data = self.db_manager.get_attendance_by_date(date)
                    for record in updated_data:
                        if record['employee_id'] == self.employee_id:
                            attendance_id = record['id']
                            break
            else:
                attendance_id = employee_attendance['id']
            
            # Add violation
            self.db_manager.add_violation(attendance_id, start_time_str, end_time_str, description)
            QMessageBox.information(self, "Sukses", "Pelanggaran berhasil ditambahkan!")
            self.accept()
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal menyimpan pelanggaran: {str(e)}")


class EditLeaveDialog(QDialog):
    """Dialog untuk edit izin"""
    def __init__(self, db_manager, leave_data, parent=None):
        super().__init__(parent)
        self.db_manager = db_manager
        self.leave_data = leave_data
        self.setWindowTitle("✏️ Edit Izin")
        self.setModal(True)
        self.resize(400, 250)
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout()
        
        # Header
        header = QLabel("✏️ EDIT IZIN")
        header.setAlignment(Qt.AlignCenter)
        header.setStyleSheet("""
            QLabel {
                font-size: 14px;
                font-weight: bold;
                padding: 15px;
                background-color: #ffc107;
                color: black;
                border-radius: 8px;
                margin-bottom: 20px;
            }
        """)
        layout.addWidget(header)
        
        # Form
        form_layout = QFormLayout()
        
        # Date selection
        self.date_edit = IndonesianDateEdit()
        from datetime import datetime
        date_obj = datetime.strptime(self.leave_data['date'], '%Y-%m-%d').date()
        self.date_edit.setDate(QDate(date_obj))
        form_layout.addRow("Tanggal Izin:", self.date_edit)
        
        # Description
        self.description_edit = QTextEdit()
        self.description_edit.setMaximumHeight(80)
        self.description_edit.setPlainText(self.leave_data['description'])
        form_layout.addRow("Keterangan:", self.description_edit)
        
        layout.addLayout(form_layout)
        
        # Buttons
        button_layout = QHBoxLayout()
        
        save_btn = QPushButton("💾 Simpan")
        save_btn.clicked.connect(self.save_leave)
        
        cancel_btn = QPushButton("❌ Batal")
        cancel_btn.clicked.connect(self.reject)
        
        button_layout.addWidget(save_btn)
        button_layout.addWidget(cancel_btn)
        layout.addLayout(button_layout)
        
        self.setLayout(layout)
    
    def save_leave(self):
        """Simpan perubahan izin"""
        if not self.description_edit.toPlainText().strip():
            QMessageBox.warning(self, "Error", "Keterangan izin tidak boleh kosong!")
            return
        
        date = self.date_edit.date().toPython().strftime('%Y-%m-%d')
        description = self.description_edit.toPlainText().strip()
        
        try:
            self.db_manager.update_leave(self.leave_data['id'], date, description)
            QMessageBox.information(self, "Sukses", "Izin berhasil diperbarui!")
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal memperbarui izin: {str(e)}")


if __name__ == "__main__":
    main()
